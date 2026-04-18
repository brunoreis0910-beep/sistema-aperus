from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .permissions import IsStaffOrHasPermission
from django.http import HttpResponse
from datetime import datetime
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import logging
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


class RelatoriosViewSet(viewsets.ViewSet):
    """
    ViewSet para geração de relatórios em PDF
    """
    permission_classes = [IsStaffOrHasPermission]
    permission_required = 'relatorios_acessar'
    
    @action(detail=False, methods=['post'], url_path='gerar')
    def gerar_relatorio(self, request):
        """
        Gera relatório em PDF ou Excel com base nos filtros fornecidos
        """
        tipo = request.data.get('tipo')
        filtros = request.data.get('filtros', {})
        formato = filtros.get('formato', 'pdf')  # pdf ou excel
        
        # Mapear tipo de relatório para função geradora
        if formato == 'excel':
            geradores = {
                'vendas': self.gerar_relatorio_vendas_excel,
                'estoque': self.gerar_relatorio_estoque_excel,
                'compras': self.gerar_relatorio_compras_excel,
                'financeiro': self.gerar_relatorio_financeiro_excel,
                'clientes': self.gerar_relatorio_clientes_excel,
                'produtos': self.gerar_relatorio_produtos_excel,
                'desempenho': self.gerar_relatorio_desempenho_excel,
                'consolidado': self.gerar_relatorio_consolidado_excel,
                'comissoes': self.gerar_relatorio_comissoes_excel,
                'devolucoes': self.gerar_relatorio_devolucoes_excel,
                'trocas': self.gerar_relatorio_trocas_excel,
                'comandas': self.gerar_relatorio_comandas_excel,
                'cte': self.gerar_relatorio_cte_excel,
                'mdfe': self.gerar_relatorio_mdfe_excel,
                'dre': self.gerar_relatorio_dre,
            }
        else:
            geradores = {
                'vendas': self.gerar_relatorio_vendas,
                'estoque': self.gerar_relatorio_estoque,
                'compras': self.gerar_relatorio_compras,
                'financeiro': self.gerar_relatorio_financeiro,
                'clientes': self.gerar_relatorio_clientes,
                'produtos': self.gerar_relatorio_produtos,
                'desempenho': self.gerar_relatorio_desempenho,
                'consolidado': self.gerar_relatorio_consolidado,
                'comissoes': self.gerar_relatorio_comissoes,
                'devolucoes': self.gerar_relatorio_devolucoes,
                'trocas': self.gerar_relatorio_trocas,
                'comandas': self.gerar_relatorio_comandas,
                'cte': self.gerar_relatorio_cte,
                'mdfe': self.gerar_relatorio_mdfe,
                'dre': self.gerar_relatorio_dre,
            }
        
        gerador = geradores.get(tipo)
        if not gerador:
            return Response(
                {'error': f'Tipo de relatório "{tipo}" não encontrado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            return gerador(filtros)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            logger.error(f'Erro ao gerar relatório {tipo}: {e}')
            logger.error(f'Traceback: {error_trace}')
            print(f'Erro completo ao gerar relatório {tipo}:')
            print(error_trace)
            return Response(
                {'error': f'Erro ao gerar relatório: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def criar_pdf_base(self, titulo, descricao=''):
        """Cria estrutura base do PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(titulo, title_style))
        
        # Descrição
        if descricao:
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(descricao, subtitle_style))
        
        # Data de geração
        data_geracao = datetime.now().strftime('%d/%m/%Y %H:%M')
        elements.append(Paragraph(f'Gerado em: {data_geracao}', subtitle_style))
        elements.append(Spacer(1, 0.5*cm))
        
        return buffer, doc, elements, styles
    

    def _adicionar_tabela_resumo(self, elementos, titulo, dados):
        """Helper para adicionar tabela de resumo"""
        from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet
        
        if not dados:
            return

        styles = getSampleStyleSheet()
        elementos.append(Paragraph(f'<b>{titulo}</b>', styles['Heading2']))
        
        # Dados com cabeçalho
        tabela_dados = [['Descrição', 'Total (R$)', '%']]
        total_geral = sum(float(row[1] or 0) for row in dados)
        
        for nome, valor in dados:
            valor = float(valor or 0)
            percentual = (valor / total_geral * 100) if total_geral > 0 else 0
            
            valor_fmt = f'R$ {valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            perc_fmt = f'{percentual:.1f}%'.replace('.', ',')
            
            tabela_dados.append([nome, valor_fmt, perc_fmt])
            
        tabela_dados.append(['TOTAL', f'R$ {total_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'), '100%'])

        table = Table(tabela_dados, colWidths=[10*cm, 4*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')), # Total Row
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ]))
        
        elementos.append(table)
        elementos.append(Spacer(1, 0.5*cm))

    def gerar_relatorio_vendas(self, filtros):
        """Gera relatório de vendas"""
        from .models import Cliente, Produto
        from django.db import connection
        from datetime import datetime
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Vendas',
            'Análise completa de vendas no período selecionado'
        )
        
        # Mostrar filtros aplicados com nomes reais
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('cliente'):
            try:
                cliente = Cliente.objects.get(id_cliente=filtros['cliente'])
                filtros_text.append(f"Cliente: {cliente.nome_razao_social}")
            except Cliente.DoesNotExist:
                filtros_text.append(f"Cliente ID: {filtros['cliente']}")
        if filtros.get('produto'):
            try:
                produto = Produto.objects.get(id_produto=filtros['produto'])
                filtros_text.append(f"Produto: {produto.nome_produto}")
            except Produto.DoesNotExist:
                filtros_text.append(f"Produto ID: {filtros['produto']}")
        if filtros.get('status') and filtros['status'] != 'todos':
            filtros_text.append(f"Status: {filtros['status']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # 1. LISTAGEM DE VENDAS
        elements.append(Paragraph('Listagem de Vendas (Últimas 100)', styles['Heading2']))

        # Buscar vendas reais do banco de dados usando SQL direto
        # Calcula o valor total somando os itens da venda
        sql = """
            SELECT 
                v.id_venda,
                v.data_documento,
                v.numero_documento,
                c.nome_razao_social,
                COALESCE(
                    (SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda),
                    v.valor_total_venda,
                    v.valor_total,
                    0
                ) as valor,
                v.status_venda
            FROM vendas v
            LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
            WHERE 1=1
        """
        where_clause = ""
        params = []
        
        # Aplicar filtros
        if filtros.get('dataInicio'):
            where_clause += " AND DATE(v.data_documento) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            where_clause += " AND DATE(v.data_documento) <= %s"
            params.append(filtros['dataFim'] + ' 23:59:59')
        if filtros.get('cliente'):
            where_clause += " AND v.id_cliente = %s"
            params.append(filtros['cliente'])
        if filtros.get('status') and filtros['status'] != 'todos':
            where_clause += " AND v.status_venda = %s"
            params.append(filtros['status'])
        
        sql_completo = sql + where_clause + " ORDER BY v.data_documento DESC LIMIT 100"
        
        # Executar query listagem
        with connection.cursor() as cursor:
            cursor.execute(sql_completo, params)
            vendas = cursor.fetchall()
        
        # Criar tabela com dados reais
        data = [['#', 'Documento', 'Data', 'Cliente', 'Valor', 'Status']]
        total_vendas = 0
        
        for idx, venda in enumerate(vendas, 1):
            id_venda, data_venda, numero_doc, cliente_nome, valor, status = venda
            data_formatada = data_venda.strftime('%d/%m/%Y') if data_venda else '-'
            numero_doc = numero_doc or str(id_venda)
            cliente_nome = cliente_nome or 'N/A'
            valor = float(valor) if valor else 0
            valor_formatado = f'R$ {valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            status = status or 'Pendente'
            
            data.append([str(idx), numero_doc, data_formatada, cliente_nome, valor_formatado, status])
            total_vendas += valor
        
        if len(data) == 1:  # Apenas cabeçalho
            data.append(['-', '-', '-', 'Nenhuma venda encontrada', '-', '-'])
        
        table = Table(data, colWidths=[1*cm, 2*cm, 2.5*cm, 6*cm, 3*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo da tabela
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            
            # Alinhamentos específicos
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # # - Centro
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Documento - Centro
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Data - Centro
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Cliente - Esquerda
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Valor - Direita
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Status - Centro
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 1*cm))

        # Adicionar Resumos Solicitados
        elements.append(Paragraph('Resumos Analíticos', styles['Heading1']))
        elements.append(Spacer(1, 0.5*cm))

        with connection.cursor() as cursor:
            # 1. Resumo por Tipo de Pagamento
            sql_pag = f"""
                SELECT 
                    COALESCE(fc.forma_pagamento, 'Diversos/Não Identificado'), 
                    SUM(fc.valor_parcela)
                FROM financeiro_contas fc
                JOIN vendas v ON fc.id_venda_origem = v.id_venda
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente 
                WHERE 1=1 AND fc.tipo_conta = 'Receber' {where_clause}
                GROUP BY fc.forma_pagamento
                ORDER BY 2 DESC
            """
            cursor.execute(sql_pag, params)
            self._adicionar_tabela_resumo(elements, 'Por Tipo de Pagamento', cursor.fetchall())

            # 2. Resumo por Tipo de Curso (Grupo de Produto)
            sql_grupo = f"""
                SELECT 
                    COALESCE(gp.nome_grupo, 'Sem Grupo / Curso'), 
                    SUM(vi.valor_total)
                FROM venda_itens vi
                JOIN vendas v ON vi.id_venda = v.id_venda
                JOIN produtos p ON vi.id_produto = p.id_produto
                LEFT JOIN grupos_produto gp ON p.id_grupo = gp.id_grupo
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
                WHERE 1=1 {where_clause}
                GROUP BY gp.nome_grupo
                ORDER BY 2 DESC
            """
            cursor.execute(sql_grupo, params)
            self._adicionar_tabela_resumo(elements, 'Por Tipo de Curso (Grupo do Produto)', cursor.fetchall())

            # 3. Resumo por Operação
            sql_operacao = f"""
                SELECT 
                    COALESCE(o.nome_operacao, 'Sem Operação'), 
                    SUM(v.valor_total)
                FROM vendas v
                LEFT JOIN operacoes o ON v.id_operacao = o.id_operacao
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
                WHERE 1=1 {where_clause}
                GROUP BY o.nome_operacao
                ORDER BY 2 DESC
            """
            cursor.execute(sql_operacao, params)
            self._adicionar_tabela_resumo(elements, 'Por Natureza da Operação', cursor.fetchall())
            
            # 4. Resumo por Conta de Lançamento
            sql_conta = f"""
                SELECT 
                    COALESCE(cb.nome_conta, 'Não Bancário/Caixa Geral'), 
                    SUM(fc.valor_parcela)
                FROM financeiro_contas fc
                LEFT JOIN contas_bancarias cb ON fc.id_conta_baixa = cb.id_conta_bancaria
                JOIN vendas v ON fc.id_venda_origem = v.id_venda
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
                WHERE 1=1 AND fc.tipo_conta = 'Receber' {where_clause}
                GROUP BY cb.nome_conta
                ORDER BY 2 DESC
            """
            cursor.execute(sql_conta, params)
            self._adicionar_tabela_resumo(elements, 'Por Conta de Lançamento', cursor.fetchall())

            # 5. Resumo por Cidade
            sql_cidade = f"""
                SELECT 
                    COALESCE(c.cidade, 'Não Informada'), 
                    SUM(v.valor_total)
                FROM vendas v
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
                WHERE 1=1 {where_clause}
                GROUP BY c.cidade
                ORDER BY 2 DESC
            """
            cursor.execute(sql_cidade, params)
            self._adicionar_tabela_resumo(elements, 'Por Cidade do Cliente', cursor.fetchall())

            # 6. Top 10 Clientes (Mudar para resumo se precisar de todos)
            sql_cliente = f"""
                SELECT 
                    COALESCE(c.nome_razao_social, 'Consumidor Geral'), 
                    SUM(v.valor_total)
                FROM vendas v
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
                WHERE 1=1 {where_clause}
                GROUP BY c.nome_razao_social
                ORDER BY 2 DESC
                LIMIT 10
            """
            cursor.execute(sql_cliente, params)
            self._adicionar_tabela_resumo(elements, 'Top 10 Clientes', cursor.fetchall())
        
        # Mostrar total real
        # Recalcular total geral baseado na query principal (sem limite)
        with connection.cursor() as cursor:
            sql_total = f"""
                SELECT SUM(v.valor_total) 
                FROM vendas v 
                LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
                WHERE 1=1 {where_clause}
            """
            cursor.execute(sql_total, params)
            total_geral_real = cursor.fetchone()[0] or 0

        total_formatado = f'R$ {total_geral_real:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        elements.append(Paragraph(f'<b>Total Geral de Vendas:</b> {total_formatado}', styles['Heading3']))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_vendas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_financeiro(self, filtros):
        """Gera relatório financeiro com dados de vendas e compras"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório Financeiro',
            'Fluxo de caixa - Entradas e Saídas'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        
        tipo = filtros.get('tipo', 'todos')  # 'entrada', 'saida', 'todos'
        if tipo == 'entrada':
            filtros_text.append("Tipo: Entradas (Vendas)")
        elif tipo == 'saida':
            filtros_text.append("Tipo: Saídas (Compras)")
        else:
            filtros_text.append("Tipo: Todos")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Debug: verificar tabelas disponíveis
        with connection.cursor() as cursor_debug:
            cursor_debug.execute("SHOW TABLES")
            tables = cursor_debug.fetchall()
            print(f"Tabelas disponíveis: {[t[0] for t in tables]}")
        
        # Buscar movimentações financeiras (vendas e compras)
        movimentacoes = []
        
        # ENTRADAS - Vendas
        if tipo in ['entrada', 'todos']:
            sql_vendas = """
                SELECT 
                    v.id_venda,
                    v.data_documento,
                    CONCAT('Venda #', COALESCE(v.numero_documento, v.id_venda)) as descricao,
                    'Entrada' as tipo,
                    COALESCE((SELECT SUM(vi.valor_total) 
                             FROM venda_itens vi 
                             WHERE vi.id_venda = v.id_venda), 0) as valor,
                    v.status_venda
                FROM vendas v
                WHERE 1=1
            """
            params_vendas = []
            
            if filtros.get('dataInicio'):
                sql_vendas += " AND DATE(v.data_documento) >= %s"
                params_vendas.append(filtros['dataInicio'])
            
            if filtros.get('dataFim'):
                sql_vendas += " AND DATE(v.data_documento) <= %s"
                params_vendas.append(filtros['dataFim'])
            
            sql_vendas += " ORDER BY v.data_documento DESC"
            
            try:
                with connection.cursor() as cursor:
                    cursor.execute(sql_vendas, params_vendas)
                    vendas = cursor.fetchall()
                    for venda in vendas:
                        movimentacoes.append(venda)
            except Exception as e:
                print(f"Erro ao buscar vendas: {e}")
        
        # SAÍDAS - Compras
        if tipo in ['saida', 'todos']:
            sql_compras = """
                SELECT 
                    c.id_compra,
                    c.data_movimento_entrada,
                    CONCAT('Compra NF ', c.numero_nota) as descricao,
                    'Saída' as tipo,
                    COALESCE(c.valor_total_nota, 0) as valor,
                    IF(c.gerou_financeiro = 1, 'Concluída', 'Pendente') as status_compra
                FROM compras c
                WHERE 1=1
            """
            params_compras = []
            
            if filtros.get('dataInicio'):
                sql_compras += " AND DATE(c.data_movimento_entrada) >= %s"
                params_compras.append(filtros['dataInicio'])
            
            if filtros.get('dataFim'):
                sql_compras += " AND DATE(c.data_movimento_entrada) <= %s"
                params_compras.append(filtros['dataFim'])
            
            sql_compras += " ORDER BY c.data_movimento_entrada DESC"
            
            try:
                with connection.cursor() as cursor:
                    cursor.execute(sql_compras, params_compras)
                    compras = cursor.fetchall()
                    for compra in compras:
                        movimentacoes.append(compra)
            except Exception as e:
                print(f"Erro ao buscar compras: {e}")
        
        # Ordenar por data
        movimentacoes.sort(key=lambda x: x[1] if x[1] else '', reverse=True)
        
        # Criar tabela
        data = [['#', 'Data', 'Descrição', 'Tipo', 'Valor', 'Status']]
        col_widths = [0.8*cm, 2.2*cm, 6.5*cm, 2*cm, 2.5*cm, 2.5*cm]
        
        total_entradas = 0
        total_saidas = 0
        
        for idx, mov in enumerate(movimentacoes, 1):
            id_mov, data_mov, descricao, tipo_mov, valor, status_mov = mov
            
            # Formatar data
            if data_mov:
                if isinstance(data_mov, str):
                    try:
                        data_obj = datetime.strptime(data_mov, '%Y-%m-%d')
                        data_fmt = data_obj.strftime('%d/%m/%Y')
                    except:
                        data_fmt = data_mov
                else:
                    data_fmt = data_mov.strftime('%d/%m/%Y')
            else:
                data_fmt = '-'
            
            # Formatar valor
            valor_fmt = f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Acumular totais
            if tipo_mov == 'Entrada':
                total_entradas += float(valor)
            else:
                total_saidas += float(valor)
            
            # Truncar descrição
            descricao_fmt = descricao[:45] + '...' if len(descricao) > 45 else descricao
            
            data.append([
                str(idx),
                data_fmt,
                descricao_fmt,
                tipo_mov,
                valor_fmt,
                status_mov
            ])
        
        # Estilo da tabela
        table = Table(data, colWidths=col_widths)
        
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9c27b0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Colorir valores baseado no tipo
        for idx in range(1, len(data)):
            tipo_mov = data[idx][3]
            if tipo_mov == 'Entrada':
                style_commands.append(('TEXTCOLOR', (4, idx), (4, idx), colors.HexColor('#2e7d32')))  # Verde
                style_commands.append(('FONTNAME', (4, idx), (4, idx), 'Helvetica-Bold'))
            else:
                style_commands.append(('TEXTCOLOR', (4, idx), (4, idx), colors.HexColor('#d32f2f')))  # Vermelho
                style_commands.append(('FONTNAME', (4, idx), (4, idx), 'Helvetica-Bold'))
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Resumo financeiro
        elements.append(Spacer(1, 0.8*cm))
        
        saldo = total_entradas - total_saidas
        saldo_cor = '#2e7d32' if saldo >= 0 else '#d32f2f'
        
        resumo_style = ParagraphStyle(
            'ResumoFinanceiro',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=6
        )
        
        elements.append(Paragraph(
            f'<b>📊 Resumo Financeiro</b>', 
            resumo_style
        ))
        elements.append(Paragraph(
            f'<b>💰 Total Entradas (Vendas):</b> <font color="#2e7d32">R$ {total_entradas:,.2f}</font>'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        elements.append(Paragraph(
            f'<b>💸 Total Saídas (Compras):</b> <font color="#d32f2f">R$ {total_saidas:,.2f}</font>'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        elements.append(Paragraph(
            f'<b>💵 Saldo (Entradas - Saídas):</b> <font color="{saldo_cor}">R$ {saldo:,.2f}</font>'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))

        # --- RESUMOS FINANCEIROS ANALÍTICOS (Dados Financeiros Reais) ---
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph('Resumos Analíticos (Financeiro Real)', styles['Heading2']))
        elements.append(Paragraph('Obs: Baseado nos lançamentos financeiros (Pagar/Receber), podendo diferir do operacional acima.', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        # Filtros para financeiro_contas
        where_fin = "WHERE 1=1"
        params_fin = []
        if filtros.get('dataInicio'):
            where_fin += " AND data_vencimento >= %s"
            params_fin.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            where_fin += " AND data_vencimento <= %s"
            params_fin.append(filtros['dataFim'])

        try:
            with connection.cursor() as cursor:
                # 1. Resumo por Centro de Custo
                sql_cc = f"""
                    SELECT 
                        COALESCE(cc.nome_centro_custo, 'Sem Centro de Custo'),
                        SUM(fc.valor_parcela) as total
                    FROM financeiro_contas fc
                    LEFT JOIN centro_custo cc ON fc.id_centro_custo = cc.id_centro_custo
                    {where_fin}
                    GROUP BY cc.nome_centro_custo
                    ORDER BY total DESC
                """
                cursor.execute(sql_cc, params_fin)
                res_cc = cursor.fetchall()
                if res_cc:
                    self._adicionar_tabela_resumo(elements, 'Por Centro de Custo', res_cc)

                # 2. Resumo por Conta Bancária
                sql_cb = f"""
                    SELECT 
                        COALESCE(cb.nome_conta, 'Sem Conta Vinculada'),
                        SUM(fc.valor_parcela) as total
                    FROM financeiro_contas fc
                    LEFT JOIN contas_bancarias cb ON fc.id_conta_baixa = cb.id_conta_bancaria
                    {where_fin} AND fc.status_conta IN ('Paga', 'Pago', 'Baixada')
                    GROUP BY cb.nome_conta
                    ORDER BY total DESC
                """
                cursor.execute(sql_cb, params_fin)
                res_cb = cursor.fetchall()
                if res_cb:
                    self._adicionar_tabela_resumo(elements, 'Por Conta Bancária (Baixados)', res_cb)

                # 3. Resumo por Forma de Pagamento
                sql_fp = f"""
                    SELECT 
                        COALESCE(fc.forma_pagamento, 'Diversos'),
                        SUM(fc.valor_parcela) as total
                    FROM financeiro_contas fc
                    {where_fin}
                    GROUP BY fc.forma_pagamento
                    ORDER BY total DESC
                """
                cursor.execute(sql_fp, params_fin)
                res_fp = cursor.fetchall()
                if res_fp:
                    self._adicionar_tabela_resumo(elements, 'Por Forma de Pagamento', res_fp)
        except Exception as e:
            print(f"Erro ao gerar resumos analíticos financeiros: {e}")
            elements.append(Paragraph(f"Erro ao gerar resumos analíticos: {str(e)}", styles['Normal']))

        if not movimentacoes:
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph(
                '<i>Nenhuma movimentação encontrada para o período selecionado.</i>',
                styles['Normal']
            ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_financeiro_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_dre(self, filtros):
        """Gera relatório DRE em PDF"""
        from decimal import Decimal
        import datetime as dt
        import calendar
        from django.utils import timezone
        from django.db.models import Sum, Q, Subquery, OuterRef, DecimalField
        from django.db.models.functions import Coalesce
        from .models import Venda, VendaItem, FinanceiroConta, ContaServico, Estoque

        def _dec(v):
            return Decimal('0') if v is None else Decimal(str(v))

        def _fmt_reais(v):
            return f"R$ {float(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        def _pct(v):
            return f"{float(v):.2f}%"

        # --- período ---
        data_inicio_str = filtros.get('dataInicio')
        data_fim_str = filtros.get('dataFim')
        hoje = dt.date.today()

        if data_inicio_str and data_fim_str:
            try:
                di = dt.date.fromisoformat(data_inicio_str)
                df = dt.date.fromisoformat(data_fim_str)
            except ValueError:
                di = dt.date(hoje.year, hoje.month, 1)
                df = dt.date(hoje.year, hoje.month, calendar.monthrange(hoje.year, hoje.month)[1])
        else:
            di = dt.date(hoje.year, hoje.month, 1)
            df = dt.date(hoje.year, hoje.month, calendar.monthrange(hoje.year, hoje.month)[1])

        di_dt = timezone.make_aware(dt.datetime.combine(di, dt.time.min))
        df_dt = timezone.make_aware(dt.datetime.combine(df, dt.time(23, 59, 59)))

        # --- vendas ---
        vendas_qs = Venda.objects.filter(
            data_documento__gte=di_dt,
            data_documento__lte=df_dt,
        ).exclude(status_nfe='CANCELADA')
        receita_bruta = _dec(vendas_qs.aggregate(t=Sum('valor_total'))['t'])

        itens_qs = VendaItem.objects.filter(id_venda__in=vendas_qs)
        impostos_agg = itens_qs.aggregate(
            icms=Coalesce(Sum('valor_icms'), Decimal('0'), output_field=DecimalField()),
            icms_st=Coalesce(Sum('valor_icms_st'), Decimal('0'), output_field=DecimalField()),
            pis=Coalesce(Sum('valor_pis'), Decimal('0'), output_field=DecimalField()),
            cofins=Coalesce(Sum('valor_cofins'), Decimal('0'), output_field=DecimalField()),
            ipi=Coalesce(Sum('valor_ipi'), Decimal('0'), output_field=DecimalField()),
        )
        total_impostos = sum(_dec(impostos_agg[k]) for k in ['icms', 'icms_st', 'pis', 'cofins', 'ipi'])
        receita_liquida = receita_bruta - total_impostos

        # --- CMV ---
        estoque_custo = Estoque.objects.filter(id_produto=OuterRef('id_produto')).values('custo_medio')[:1]
        itens_com_custo = itens_qs.annotate(
            custo_unit=Coalesce(Subquery(estoque_custo, output_field=DecimalField()), Decimal('0'), output_field=DecimalField())
        )
        cmv = sum(_dec(it['quantidade']) * _dec(it['custo_unit']) for it in itens_com_custo.values('quantidade', 'custo_unit'))
        resultado_bruto = receita_liquida - cmv

        # --- despesas ---
        despesas_qs = FinanceiroConta.objects.filter(
            tipo_conta='pagar', status_conta='Pago',
            data_pagamento__gte=di, data_pagamento__lte=df,
            id_compra_origem__isnull=True,
        )
        total_despesas = _dec(despesas_qs.aggregate(t=Sum('valor_parcela'))['t'])
        despesas_por_depto = list(
            despesas_qs.values('id_departamento__nome_departamento')
            .annotate(total=Coalesce(Sum('valor_parcela'), Decimal('0'), output_field=DecimalField()))
            .order_by('-total')
        )

        # --- serviços ---
        servicos_qs = ContaServico.objects.filter(status='pago', data_pagamento__gte=di, data_pagamento__lte=df)
        total_servicos = _dec(servicos_qs.aggregate(t=Sum('valor_total'))['t'])

        # --- receitas nao-op ---
        rec_extra = _dec(FinanceiroConta.objects.filter(
            tipo_conta='receber', status_conta='Pago',
            data_pagamento__gte=di, data_pagamento__lte=df,
            id_venda_origem__isnull=True, id_os_origem__isnull=True,
        ).aggregate(t=Sum('valor_parcela'))['t'])

        total_despesas_all = total_despesas + total_servicos
        resultado_operacional = resultado_bruto - total_despesas_all + rec_extra
        margem_bruta = (resultado_bruto / receita_bruta * 100) if receita_bruta else Decimal('0')
        margem_op = (resultado_operacional / receita_bruta * 100) if receita_bruta else Decimal('0')

        # --- construir PDF ---
        periodo_label = f"{di.strftime('%d/%m/%Y')} a {df.strftime('%d/%m/%Y')}"
        buffer, doc, elements, styles = self.criar_pdf_base(
            'DRE — Demonstração do Resultado',
            f'Período: {periodo_label}'
        )

        label_style = ParagraphStyle('DRELabel', parent=styles['Normal'], fontSize=9, leading=13)
        dest_style = ParagraphStyle('DREDest', parent=styles['Normal'], fontSize=9, leading=13, textColor=colors.HexColor('#7f8c8d'))
        bold_style = ParagraphStyle('DREBold', parent=styles['Normal'], fontSize=10, leading=14, fontName='Helvetica-Bold')

        COR_VERDE = colors.HexColor('#2e7d32')
        COR_VERMELHO = colors.HexColor('#c62828')
        COR_HEADER = colors.HexColor('#004d40')
        COR_DEST = colors.HexColor('#e8f5e9')
        COR_DEST2 = colors.HexColor('#e3f2fd')

        def linha(label, valor, nivel=0, negativo=False, cor_linha=None):
            pad = '    ' * nivel
            valor_fmt = _fmt_reais(abs(valor))
            if negativo:
                valor_fmt = f'({valor_fmt})'
            cor_val = COR_VERMELHO if negativo else COR_VERDE
            return [Paragraph(f'{pad}{label}', label_style), Paragraph(valor_fmt, label_style), cor_val, cor_linha]

        # Tabela DRE
        rows = []

        # Bloco header
        def header(texto):
            return [Paragraph(f'<b>{texto}</b>', bold_style), Paragraph('', bold_style), None, COR_HEADER]

        rows.append(header('RECEITA'))
        rows.append(linha('Receita Bruta de Vendas', receita_bruta))
        rows.append(linha('(-) Impostos s/ Vendas (ICMS, PIS, COFINS, IPI)', total_impostos, nivel=1, negativo=True))
        rows.append(linha('= Receita Líquida', receita_liquida, cor_linha=COR_DEST))

        rows.append(header('CUSTO'))
        rows.append(linha('(-) CMV — Custo da Mercadoria Vendida', cmv, nivel=1, negativo=True))
        rows.append(linha('= Resultado Bruto (Lucro Bruto)', resultado_bruto, cor_linha=COR_DEST))
        rows.append(linha(f'   Margem Bruta', margem_bruta, nivel=1))

        rows.append(header('DESPESAS OPERACIONAIS'))
        if despesas_por_depto:
            for d in despesas_por_depto:
                nome = d['id_departamento__nome_departamento'] or 'Sem Departamento'
                rows.append(linha(f'(-) {nome}', _dec(d['total']), nivel=1, negativo=True))
        else:
            rows.append([Paragraph('    Nenhuma despesa no período', dest_style), Paragraph('R$ 0,00', dest_style), None, None])

        if total_servicos > 0:
            rows.append(linha('(-) Contas de Serviços / Utilidades', total_servicos, nivel=1, negativo=True))

        if rec_extra > 0:
            rows.append(header('RECEITAS NÃO OPERACIONAIS'))
            rows.append(linha('(+) Outras Receitas', rec_extra, nivel=1))

        rows.append(linha('= Resultado Operacional (EBIT)', resultado_operacional, cor_linha=COR_DEST2))
        rows.append(linha(f'   Margem Operacional', margem_op, nivel=1))

        col_w = [11 * cm, 5 * cm]
        tbl_data = [[Paragraph('<b>Descrição</b>', bold_style), Paragraph('<b>Valor</b>', bold_style)]]
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), COR_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

        for row_idx, (label_p, val_p, cor_val, cor_bg) in enumerate(rows, start=1):
            tbl_data.append([label_p, val_p])
            if cor_bg:
                style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), cor_bg))
                style_cmds.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white))
            elif cor_val:
                style_cmds.append(('TEXTCOLOR', (1, row_idx), (1, row_idx), cor_val))
                style_cmds.append(('FONTNAME', (1, row_idx), (1, row_idx), 'Helvetica-Bold'))

        tbl = Table(tbl_data, colWidths=col_w)
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)

        doc.build(elements)
        buffer.seek(0)
        filename = f'dre_{di.strftime("%Y%m")}_{datetime.now().strftime("%H%M%S")}.pdf'
        resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    def gerar_relatorio_estoque(self, filtros):
        """Gera relatório de estoque com valores"""
        from .models import Produto
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Estoque',
            'Posição atual de estoque por produto e depósito com valores'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('produto'):
            try:
                produto = Produto.objects.get(id_produto=filtros['produto'])
                filtros_text.append(f"Produto: {produto.nome_produto}")
            except Produto.DoesNotExist:
                filtros_text.append(f"Produto ID: {filtros['produto']}")
        
        if filtros.get('deposito'):
            filtros_text.append(f"Depósito ID: {filtros['deposito']}")
        
        tipo_valor = filtros.get('tipoValor', 'todos')  # 'venda', 'compra', 'custo', 'todos'
        if tipo_valor == 'venda':
            filtros_text.append("Mostrando: Valor de Venda")
        elif tipo_valor == 'compra':
            filtros_text.append("Mostrando: Valor de Compra")
        elif tipo_valor == 'custo':
            filtros_text.append("Mostrando: Valor de Custo")
        else:
            filtros_text.append("Mostrando: Todos os Valores")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar estoque real do banco com valores (da tabela estoque)
        sql = """
            SELECT 
                p.id_produto,
                p.codigo_produto,
                p.nome_produto,
                CAST(COALESCE(e.id_deposito, 0) AS CHAR) as deposito_info,
                CAST(COALESCE(e.quantidade, 0) AS DECIMAL(10,2)) as quantidade,
                CAST(COALESCE(e.valor_venda, 0) AS DECIMAL(10,2)) as valor_venda,
                CAST(COALESCE(e.valor_ultima_compra, 0) AS DECIMAL(10,2)) as valor_compra,
                CAST(COALESCE(e.valor_ultima_compra, 0) AS DECIMAL(10,2)) as preco_custo
            FROM produtos p
            LEFT JOIN estoque e ON p.id_produto = e.id_produto
            WHERE 1=1
        """
        params = []
        
        if filtros.get('produto'):
            sql += " AND p.id_produto = %s"
            params.append(filtros['produto'])
        
        if filtros.get('deposito'):
            sql += " AND e.id_deposito = %s"
            params.append(filtros['deposito'])
        
        sql += " ORDER BY p.nome_produto"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                estoques = cursor.fetchall()
        except Exception as e:
            print(f"Erro na query de estoque: {e}")
            raise
        
        # Criar cabeçalho da tabela baseado no tipo de valor
        if tipo_valor == 'venda':
            data = [['#', 'Código', 'Produto', 'Depósito', 'Qtd', 'Preço Unit.', 'Valor Total']]
            col_widths = [0.8*cm, 2.2*cm, 6*cm, 2.3*cm, 1.2*cm, 2.2*cm, 2.3*cm]
        elif tipo_valor == 'compra':
            data = [['#', 'Código', 'Produto', 'Depósito', 'Qtd', 'Preço Unit.', 'Valor Total']]
            col_widths = [0.8*cm, 2.2*cm, 6*cm, 2.3*cm, 1.2*cm, 2.2*cm, 2.3*cm]
        elif tipo_valor == 'custo':
            data = [['#', 'Código', 'Produto', 'Depósito', 'Qtd', 'Preço Unit.', 'Valor Total']]
            col_widths = [0.8*cm, 2.2*cm, 6*cm, 2.3*cm, 1.2*cm, 2.2*cm, 2.3*cm]
        else:  # todos
            data = [['#', 'Código', 'Produto', 'Dep.', 'Qtd', 'Vlr Venda', 'Vlr Compra', 'Vlr Custo']]
            col_widths = [0.7*cm, 2*cm, 5.5*cm, 1.8*cm, 1*cm, 2.1*cm, 2.1*cm, 2.1*cm]
        
        total_geral_venda = 0
        total_geral_compra = 0
        total_geral_custo = 0
        
        for idx, estoque in enumerate(estoques, 1):
            id_produto, codigo_produto, nome_produto, deposito_info, quantidade, valor_venda, valor_compra, preco_custo = estoque
            
            # Pular produtos sem estoque (quantidade = 0)
            quantidade = float(quantidade) if quantidade else 0
            if quantidade == 0:
                continue
            
            # Formatar código e depósito
            codigo_produto = str(codigo_produto) if codigo_produto else '-'
            # Limitar tamanho do código para não quebrar layout
            if len(codigo_produto) > 15:
                codigo_produto = codigo_produto[:12] + '...'
            
            nome_deposito = f"Dep. {deposito_info}" if deposito_info and deposito_info != '0' else 'Geral'
            
            valor_venda = float(valor_venda) if valor_venda else 0
            valor_compra = float(valor_compra) if valor_compra else 0
            preco_custo = float(preco_custo) if preco_custo else 0
            
            total_venda = quantidade * valor_venda
            total_compra = quantidade * valor_compra
            total_custo = quantidade * preco_custo
            
            total_geral_venda += total_venda
            total_geral_compra += total_compra
            total_geral_custo += total_custo
            
            # Renumerar baseado nos itens realmente mostrados
            row_num = len(data)
            
            if tipo_valor == 'venda':
                data.append([
                    str(row_num),
                    codigo_produto,
                    nome_produto,
                    nome_deposito,
                    f"{quantidade:.0f}",
                    f"R$ {valor_venda:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {total_venda:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
            elif tipo_valor == 'compra':
                data.append([
                    str(row_num),
                    codigo_produto,
                    nome_produto,
                    nome_deposito,
                    f"{quantidade:.0f}",
                    f"R$ {valor_compra:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {total_compra:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
            elif tipo_valor == 'custo':
                data.append([
                    str(row_num),
                    codigo_produto,
                    nome_produto,
                    nome_deposito,
                    f"{quantidade:.0f}",
                    f"R$ {preco_custo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {total_custo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
            else:  # todos
                data.append([
                    str(row_num),
                    codigo_produto,
                    nome_produto,
                    nome_deposito,
                    f"{quantidade:.0f}",
                    f"R$ {total_venda:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {total_compra:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                    f"R$ {total_custo:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
        
        if len(data) == 1:
            if tipo_valor == 'todos':
                data.append(['-', '-', 'Nenhum produto encontrado', '-', '-', '-', '-', '-'])
            else:
                data.append(['-', '-', 'Nenhum produto encontrado', '-', '-', '-', '-'])
        
        table = Table(data, colWidths=col_widths)
        
        # Estilos base
        table_style = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            
            # Corpo
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Alinhamentos
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # #
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Código
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Produto
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Depósito
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Quantidade
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Código em fonte menor se for muito longo
            ('FONTSIZE', (1, 1), (1, -1), 7),  # Código menor
        ]
        
        # Alinhamentos específicos por tipo
        if tipo_valor == 'todos':
            table_style.extend([
                ('ALIGN', (5, 1), (5, -1), 'RIGHT'),  # Vlr Venda
                ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Vlr Compra
                ('ALIGN', (7, 1), (7, -1), 'RIGHT'),  # Vlr Custo
            ])
        else:
            table_style.extend([
                ('ALIGN', (5, 1), (5, -1), 'RIGHT'),  # Vlr Unit.
                ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Vlr Total
                ('FONTNAME', (6, 1), (6, -1), 'Helvetica-Bold'),  # Valor Total em negrito
            ])
        
        table.setStyle(TableStyle(table_style))
        
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Resumo com destaque visual
        from reportlab.lib.styles import ParagraphStyle
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#2e7d32'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        total_itens = len(data) - 1
        elements.append(Paragraph(f'📊 Total de Itens em Estoque: {total_itens}', summary_style))
        
        if tipo_valor == 'venda':
            elements.append(Paragraph(
                f'💰 Valor Total em Estoque: R$ {total_geral_venda:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
        elif tipo_valor == 'compra':
            elements.append(Paragraph(
                f'💰 Valor Total em Estoque: R$ {total_geral_compra:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
        elif tipo_valor == 'custo':
            elements.append(Paragraph(
                f'💰 Valor Total em Estoque: R$ {total_geral_custo:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
        else:  # todos
            elements.append(Paragraph(
                f'💵 Valor Total (Venda): R$ {total_geral_venda:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
            elements.append(Paragraph(
                f'💸 Valor Total (Compra): R$ {total_geral_compra:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
            elements.append(Paragraph(
                f'💲 Valor Total (Custo): R$ {total_geral_custo:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))

        # --- RESUMO POR GRUPO DE PRODUTO ---
        elements.append(Spacer(1, 0.8*cm))
        try:
            with connection.cursor() as cursor:
                # Reutiliza filtros de deposito e produto
                sql_grupo = """
                    SELECT 
                        COALESCE(gp.nome_grupo, 'Sem Grupo'),
                        SUM(COALESCE(e.quantidade, 0) * COALESCE(e.valor_venda, 0)) as total_venda
                    FROM produtos p
                    LEFT JOIN estoque e ON p.id_produto = e.id_produto
                    LEFT JOIN grupos_produto gp ON p.id_grupo = gp.id_grupo
                    WHERE CAST(COALESCE(e.quantidade, 0) AS DECIMAL(10,2)) > 0
                """
                params_grupo = []
                
                if filtros.get('deposito'):
                    sql_grupo += " AND e.id_deposito = %s"
                    params_grupo.append(filtros['deposito'])
                    
                if filtros.get('produto'):
                    sql_grupo += " AND p.id_produto = %s"
                    params_grupo.append(filtros['produto'])
                    
                sql_grupo += " GROUP BY gp.nome_grupo ORDER BY total_venda DESC"
                
                cursor.execute(sql_grupo, params_grupo)
                res_grupo = cursor.fetchall()
                
                if res_grupo:
                     self._adicionar_tabela_resumo(elements, 'Valor em Estoque (PV) por Grupo', res_grupo)
        except Exception as e:
            print(f"Erro ao gerar resumo estoque: {e}")
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_compras(self, filtros):
        """Gera relatório de compras"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Compras',
            'Histórico de compras e fornecedores'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('fornecedor'):
            filtros_text.append(f"Fornecedor ID: {filtros['fornecedor']}")
        if filtros.get('produto'):
            filtros_text.append(f"Produto ID: {filtros['produto']}")
        
        status = filtros.get('status', 'todos')
        if status != 'todos':
            filtros_text.append(f"Status: {status.capitalize()}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar compras do banco com valor dos itens
        sql = """
            SELECT 
                c.id_compra,
                c.numero_nota,
                c.data_movimento_entrada,
                COALESCE(f.nome_fantasia, f.nome_razao_social) as nome_fornecedor,
                COALESCE(c.valor_total_nota, 0) as valor_total,
                IF(c.gerou_financeiro = 1, 'Concluída', 'Pendente') as status_compra
            FROM compras c
            LEFT JOIN fornecedores f ON c.id_fornecedor = f.id_fornecedor
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(c.data_movimento_entrada) >= %s"
            params.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql += " AND DATE(c.data_movimento_entrada) <= %s"
            params.append(filtros['dataFim'])
        
        if filtros.get('fornecedor'):
            sql += " AND c.id_fornecedor = %s"
            params.append(filtros['fornecedor'])
        
        if filtros.get('produto'):
            sql += " AND EXISTS (SELECT 1 FROM compra_itens ci WHERE ci.id_compra = c.id_compra AND ci.id_produto = %s)"
            params.append(filtros['produto'])
        
        sql += " ORDER BY c.data_movimento_entrada DESC, c.id_compra DESC"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                compras = cursor.fetchall()
        except Exception as e:
            print(f"Erro na query de compras: {e}")
            raise
        
        # Criar tabela
        data = [['#', 'Doc.', 'Data', 'Fornecedor', 'Valor Total', 'Status']]
        col_widths = [0.8*cm, 1.8*cm, 2.2*cm, 7*cm, 2.7*cm, 2.5*cm]
        
        total_compras = 0
        valor_total_geral = 0
        
        for idx, compra in enumerate(compras, 1):
            id_compra, numero_nota, data_compra, nome_fornecedor, valor_total, status_compra = compra
            
            total_compras += 1
            valor_total_geral += float(valor_total) if valor_total else 0
            
            # Formatar número do documento (usa numero_nota ou ID)
            numero_doc_fmt = str(numero_nota) if numero_nota else str(id_compra)
            
            # Formatar data
            if data_compra:
                data_fmt = data_compra.strftime('%d/%m/%Y') if hasattr(data_compra, 'strftime') else str(data_compra)[:10]
            else:
                data_fmt = '-'
            
            # Nome do fornecedor
            fornecedor_fmt = nome_fornecedor if nome_fornecedor else 'Sem Fornecedor'
            if len(fornecedor_fmt) > 45:
                fornecedor_fmt = fornecedor_fmt[:42] + '...'
            
            # Valor total formatado
            valor_fmt = f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_total else 'R$ 0,00'
            
            # Status
            status_fmt = status_compra if status_compra else 'Pendente'
            
            data.append([
                str(idx),
                numero_doc_fmt,
                data_fmt,
                fornecedor_fmt,
                valor_fmt,
                status_fmt
            ])
        
        if len(data) == 1:
            data.append(['-', '-', '-', 'Nenhuma compra encontrada', '-', '-'])
        
        table = Table(data, colWidths=col_widths)
        
        table_style = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ed6c02')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            
            # Corpo
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Alinhamentos
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # #
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Doc
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Data
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Fornecedor
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Valor
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Status
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),  # Valor em negrito
        ]
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Resumo
        from reportlab.lib.styles import ParagraphStyle
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#ed6c02'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph(f'📦 Total de Compras: {total_compras}', summary_style))
        elements.append(Paragraph(
            f'💰 Valor Total: R$ {valor_total_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            summary_style
        ))
        
        if total_compras > 0:
            ticket_medio = valor_total_geral / total_compras
            elements.append(Paragraph(
                f'📊 Valor Médio por Compra: R$ {ticket_medio:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
        
        # --- RESUMO POR FORNECEDOR (TOP 10) ---
        elements.append(Spacer(1, 0.8*cm))
        elements.append(Paragraph('Top 10 Fornecedores (Por Valor)', styles['Heading2']))
        
        try:
            with connection.cursor() as cursor:
                # Reutiliza filtros de datas
                sql_forn = """
                    SELECT 
                        COALESCE(c.nome_fornecedor, 'Sem Fornecedor'),
                        SUM(COALESCE(c.valor_total_nota, 0)) as total,
                        COUNT(*) as qtd
                    FROM compras c
                    WHERE 1=1
                """
                params_forn = []
                
                # Reutilizar parametros aplicados no corpo da função
                # Obs: a função original pode nao ter salvo as vars de params, entao refazemos
                if filtros.get('dataInicio'):
                    sql_forn += " AND DATE(c.data_movimento_entrada) >= %s"
                    params_forn.append(filtros['dataInicio'])
                
                if filtros.get('dataFim'):
                    sql_forn += " AND DATE(c.data_movimento_entrada) <= %s"
                    params_forn.append(filtros['dataFim'])

                if filtros.get('fornecedor'):
                    sql_forn += " AND c.id_fornecedor = %s"
                    params_forn.append(filtros['fornecedor'])
                    
                sql_forn += " GROUP BY c.nome_fornecedor ORDER BY total DESC LIMIT 10"
                
                cursor.execute(sql_forn, params_forn)
                res_forn = cursor.fetchall()
                
                if res_forn:
                     # Adaptar para o helper _adicionar_tabela_resumo que espera (nome, valor) ou (nome, valor, percentual calc interno)
                     # Mas aqui temos qtd tambem. Vamos formatar para (Nome (Qtd), Valor)
                     dados_adaptados = []
                     for row in res_forn:
                         nome, valor, qtd = row
                         dados_adaptados.append((f"{nome} ({qtd}x)", valor))
                         
                     self._adicionar_tabela_resumo(elements, 'Compras por Fornecedor', dados_adaptados)

        except Exception as e:
            print(f"Erro resumo compras: {e}")

        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_clientes(self, filtros):
        """Gera relatório de clientes"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Clientes',
            'Cadastro completo de clientes'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('cliente'):
            filtros_text.append(f"Cliente ID: {filtros['cliente']}")
        
        status = filtros.get('status', 'todos')
        if status != 'todos':
            filtros_text.append(f"Status: {status.capitalize()}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar clientes do banco com valor correto das vendas
        sql = """
            SELECT 
                c.id_cliente,
                c.nome_razao_social,
                c.nome_fantasia,
                c.cpf_cnpj,
                c.telefone,
                c.email,
                c.cidade,
                c.estado,
                COALESCE(COUNT(DISTINCT v.id_venda), 0) as total_vendas,
                COALESCE((
                    SELECT SUM(vi.valor_total)
                    FROM vendas v2
                    LEFT JOIN venda_itens vi ON v2.id_venda = vi.id_venda
                    WHERE v2.id_cliente = c.id_cliente
                ), 0) as valor_total
            FROM clientes c
            LEFT JOIN vendas v ON c.id_cliente = v.id_cliente
            WHERE 1=1
        """
        params = []
        
        if filtros.get('cliente'):
            sql += " AND c.id_cliente = %s"
            params.append(filtros['cliente'])
        
        sql += " GROUP BY c.id_cliente, c.nome_razao_social, c.nome_fantasia, c.cpf_cnpj, "
        sql += "c.telefone, c.email, c.cidade, c.estado"
        sql += " ORDER BY c.nome_razao_social"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                clientes = cursor.fetchall()
        except Exception as e:
            print(f"Erro na query de clientes: {e}")
            raise
        
        # Criar tabela
        data = [['#', 'Nome/Razão Social', 'CPF/CNPJ', 'Cidade/UF', 'Telefone', 'Vendas', 'Valor Total']]
        col_widths = [0.7*cm, 5.5*cm, 2.8*cm, 2.5*cm, 2.2*cm, 1.3*cm, 2*cm]
        
        total_clientes = 0
        total_vendas_geral = 0
        total_valor_geral = 0
        
        for idx, cliente in enumerate(clientes, 1):
            id_cliente, nome_razao, nome_fantasia, cpf_cnpj, telefone, email, cidade, estado, total_vendas, valor_total = cliente
            
            total_clientes += 1
            total_vendas_geral += int(total_vendas) if total_vendas else 0
            total_valor_geral += float(valor_total) if valor_total else 0
            
            # Nome: usa fantasia se tiver, senão razão social
            nome_exibir = nome_fantasia if nome_fantasia else nome_razao
            if len(nome_exibir) > 40:
                nome_exibir = nome_exibir[:37] + '...'
            
            # CPF/CNPJ formatado
            cpf_cnpj_fmt = cpf_cnpj if cpf_cnpj else '-'
            
            # Cidade/UF
            cidade_uf = f"{cidade}/{estado}" if cidade and estado else (cidade or estado or '-')
            if len(cidade_uf) > 20:
                cidade_uf = cidade_uf[:17] + '...'
            
            # Telefone
            telefone_fmt = telefone if telefone else '-'
            
            # Valor total formatado
            valor_fmt = f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if valor_total else 'R$ 0,00'
            
            data.append([
                str(idx),
                nome_exibir,
                cpf_cnpj_fmt,
                cidade_uf,
                telefone_fmt,
                str(int(total_vendas)) if total_vendas else '0',
                valor_fmt
            ])
        
        if len(data) == 1:
            data.append(['-', 'Nenhum cliente encontrado', '-', '-', '-', '-', '-'])
        
        table = Table(data, colWidths=col_widths)
        
        table_style = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0288d1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            
            # Corpo
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Alinhamentos
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # #
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Nome
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # CPF/CNPJ
            ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Cidade/UF
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Telefone
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Vendas
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),   # Valor Total
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (6, 1), (6, -1), 'Helvetica-Bold'),  # Valor em negrito
        ]
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Resumo
        from reportlab.lib.styles import ParagraphStyle
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#0288d1'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph(f'👥 Total de Clientes: {total_clientes}', summary_style))
        elements.append(Paragraph(f'🛒 Total de Vendas: {total_vendas_geral}', summary_style))
        elements.append(Paragraph(
            f'💰 Valor Total: R$ {total_valor_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            summary_style
        ))
        
        if total_clientes > 0 and total_vendas_geral > 0:
            ticket_medio = total_valor_geral / total_vendas_geral
            elements.append(Paragraph(
                f'📊 Ticket Médio: R$ {ticket_medio:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                summary_style
            ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_produtos(self, filtros):
        """Gera relatório de produtos com estoque e preços"""
        from .models import Produto
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Produtos',
            'Catálogo completo com preços e estoque'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('produto'):
            try:
                produto = Produto.objects.get(id_produto=filtros['produto'])
                filtros_text.append(f"Produto: {produto.nome_produto}")
            except Produto.DoesNotExist:
                filtros_text.append(f"Produto ID: {filtros['produto']}")
        
        if filtros.get('grupo'):
            filtros_text.append(f"Grupo ID: {filtros['grupo']}")
        
        status = filtros.get('status', 'todos')
        if status == 'ativo':
            filtros_text.append("Status: Ativos")
        elif status == 'inativo':
            filtros_text.append("Status: Inativos")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Debug: verificar estrutura da tabela produtos
        with connection.cursor() as cursor_debug:
            cursor_debug.execute("DESCRIBE produtos")
            cols_produtos = cursor_debug.fetchall()
            print(f"Colunas da tabela produtos: {[col[0] for col in cols_produtos]}")
        
        # Buscar produtos do banco
        sql = """
            SELECT 
                p.id_produto,
                p.codigo_produto,
                p.nome_produto,
                p.id_grupo,
                COALESCE((SELECT SUM(e.quantidade) FROM estoque e WHERE e.id_produto = p.id_produto), 0) as estoque_total,
                COALESCE((SELECT e.valor_venda FROM estoque e WHERE e.id_produto = p.id_produto LIMIT 1), 0) as valor_venda,
                COALESCE((SELECT e.valor_ultima_compra FROM estoque e WHERE e.id_produto = p.id_produto LIMIT 1), 0) as valor_compra,
                1 as ativo
            FROM produtos p
            WHERE 1=1
        """
        params = []
        
        if filtros.get('produto'):
            sql += " AND p.id_produto = %s"
            params.append(filtros['produto'])
        
        if filtros.get('grupo'):
            sql += " AND p.id_grupo = %s"
            params.append(filtros['grupo'])
        
        sql += " ORDER BY p.nome_produto"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                produtos = cursor.fetchall()
        except Exception as e:
            print(f"Erro na query de produtos: {e}")
            raise
        
        # Buscar nomes dos grupos
        grupos_dict = {}
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id_grupo, nome_grupo FROM grupos_produto")
                grupos = cursor.fetchall()
                grupos_dict = {g[0]: g[1] for g in grupos}
        except Exception as e:
            print(f"Erro ao buscar grupos: {e}")
        
        # Criar tabela
        data = [['#', 'Código', 'Produto', 'Grupo', 'Estoque', 'Vl. Venda', 'Vl. Compra', 'Status']]
        col_widths = [0.8*cm, 2*cm, 5.5*cm, 3*cm, 1.5*cm, 2*cm, 2*cm, 1.7*cm]
        
        total_produtos = 0
        total_estoque = 0
        valor_total_estoque = 0
        
        for idx, produto in enumerate(produtos, 1):
            id_prod, codigo, nome, id_grupo, estoque_qt, vl_venda, vl_compra, ativo = produto
            
            total_produtos += 1
            total_estoque += float(estoque_qt or 0)
            valor_total_estoque += float(estoque_qt or 0) * float(vl_venda or 0)
            
            # Formatar código
            codigo_fmt = str(codigo)[:15] + '...' if codigo and len(str(codigo)) > 15 else (codigo or '-')
            
            # Formatar nome
            nome_fmt = nome[:35] + '...' if nome and len(nome) > 35 else (nome or '-')
            
            # Buscar nome do grupo
            grupo_nome = grupos_dict.get(id_grupo, '-')
            grupo_fmt = grupo_nome[:20] + '...' if len(grupo_nome) > 20 else grupo_nome
            
            # Formatar estoque
            estoque_fmt = f"{float(estoque_qt):,.1f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Formatar valores
            venda_fmt = f"R$ {float(vl_venda):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            compra_fmt = f"R$ {float(vl_compra):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Status
            status_fmt = 'Ativo' if ativo else 'Inativo'
            
            data.append([
                str(idx),
                codigo_fmt,
                nome_fmt,
                grupo_fmt,
                estoque_fmt,
                venda_fmt,
                compra_fmt,
                status_fmt
            ])
        
        # Estilo da tabela
        table = Table(data, colWidths=col_widths)
        
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d32f2f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('FONTSIZE', (1, 1), (-1, -1), 7),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('ALIGN', (4, 0), (6, -1), 'RIGHT'),
            ('ALIGN', (7, 0), (7, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Destacar valores de venda em negrito
        for idx in range(1, len(data)):
            style_commands.append(('FONTNAME', (5, idx), (5, idx), 'Helvetica-Bold'))
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Resumo
        elements.append(Spacer(1, 0.8*cm))
        
        resumo_style = ParagraphStyle(
            'ResumoProdutos',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=6
        )
        
        elements.append(Paragraph(
            f'<b>📊 Resumo do Catálogo</b>', 
            resumo_style
        ))
        elements.append(Paragraph(
            f'<b>📦 Total de Produtos:</b> {total_produtos}',
            resumo_style
        ))
        elements.append(Paragraph(
            f'<b>📊 Estoque Total (Un):</b> {total_estoque:,.1f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        elements.append(Paragraph(
            f'<b>💰 Valor Total em Estoque:</b> R$ {valor_total_estoque:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        
        if not produtos:
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph(
                '<i>Nenhum produto encontrado com os filtros selecionados.</i>',
                styles['Normal']
            ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_produtos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_desempenho(self, filtros):
        """Gera relatório de análise de desempenho com indicadores e métricas"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Análise de Desempenho',
            'Indicadores de Performance e Tendências de Negócio'
        )
        
        # Mostrar período de análise
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Período de Análise:</b> {' até '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # ===== INDICADORES DE VENDAS =====
        sql_vendas = """
            SELECT 
                COUNT(DISTINCT v.id_venda) as total_vendas,
                COALESCE(SUM((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as valor_total,
                COUNT(DISTINCT v.id_cliente) as clientes_ativos,
                COALESCE(AVG((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as ticket_medio
            FROM vendas v
            WHERE 1=1
        """
        params_vendas = []
        
        if filtros.get('dataInicio'):
            sql_vendas += " AND DATE(v.data_documento) >= %s"
            params_vendas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_vendas += " AND DATE(v.data_documento) <= %s"
            params_vendas.append(filtros['dataFim'])
        
        with connection.cursor() as cursor:
            cursor.execute(sql_vendas, params_vendas)
            vendas_metrics = cursor.fetchone()
        
        total_vendas, valor_vendas, clientes_ativos, ticket_medio = vendas_metrics
        
        # ===== INDICADORES DE COMPRAS =====
        sql_compras = """
            SELECT 
                COUNT(c.id_compra) as total_compras,
                COALESCE(SUM(c.valor_total_nota), 0) as valor_total,
                COUNT(DISTINCT c.id_fornecedor) as fornecedores_ativos
            FROM compras c
            WHERE 1=1
        """
        params_compras = []
        
        if filtros.get('dataInicio'):
            sql_compras += " AND DATE(c.data_movimento_entrada) >= %s"
            params_compras.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_compras += " AND DATE(c.data_movimento_entrada) <= %s"
            params_compras.append(filtros['dataFim'])
        
        with connection.cursor() as cursor:
            cursor.execute(sql_compras, params_compras)
            compras_metrics = cursor.fetchone()
        
        total_compras, valor_compras, fornecedores_ativos = compras_metrics
        
        # ===== TOP 5 PRODUTOS MAIS VENDIDOS =====
        sql_top_produtos = """
            SELECT 
                p.nome_produto,
                SUM(vi.quantidade) as qtd_vendida,
                SUM(vi.valor_total) as valor_total
            FROM venda_itens vi
            JOIN produtos p ON vi.id_produto = p.id_produto
            JOIN vendas v ON vi.id_venda = v.id_venda
            WHERE 1=1
        """
        params_top = []
        
        if filtros.get('dataInicio'):
            sql_top_produtos += " AND DATE(v.data_documento) >= %s"
            params_top.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_top_produtos += " AND DATE(v.data_documento) <= %s"
            params_top.append(filtros['dataFim'])
        
        sql_top_produtos += " GROUP BY p.id_produto, p.nome_produto ORDER BY valor_total DESC LIMIT 5"
        
        with connection.cursor() as cursor:
            cursor.execute(sql_top_produtos, params_top)
            top_produtos = cursor.fetchall()
        
        # ===== TOP 5 CLIENTES =====
        sql_top_clientes = """
            SELECT 
                c.nome_razao_social,
                COUNT(DISTINCT v.id_venda) as total_vendas,
                COALESCE(SUM((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as valor_total
            FROM clientes c
            JOIN vendas v ON c.id_cliente = v.id_cliente
            WHERE 1=1
        """
        params_clientes = []
        
        if filtros.get('dataInicio'):
            sql_top_clientes += " AND DATE(v.data_documento) >= %s"
            params_clientes.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_top_clientes += " AND DATE(v.data_documento) <= %s"
            params_clientes.append(filtros['dataFim'])
        
        sql_top_clientes += " GROUP BY c.id_cliente, c.nome_razao_social ORDER BY valor_total DESC LIMIT 5"
        
        with connection.cursor() as cursor:
            cursor.execute(sql_top_clientes, params_clientes)
            top_clientes = cursor.fetchall()
        
        # ===== INDICADORES DE DEVOLUÇÕES =====
        sql_devolucoes = """
            SELECT 
                COUNT(d.id_devolucao) as total_devolucoes,
                COALESCE(SUM(d.valor_total_devolucao), 0) as valor_total
            FROM devolucoes d
            WHERE 1=1
        """
        params_devolucoes = []
        if filtros.get('dataInicio'):
            sql_devolucoes += " AND DATE(d.data_devolucao) >= %s"
            params_devolucoes.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_devolucoes += " AND DATE(d.data_devolucao) <= %s"
            params_devolucoes.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_devolucoes, params_devolucoes)
                devolucoes_metrics = cursor.fetchone()
            total_devolucoes, valor_devolucoes = devolucoes_metrics if devolucoes_metrics else (0, 0)
        except:
            total_devolucoes, valor_devolucoes = 0, 0
        
        # ===== INDICADORES DE TROCAS =====
        sql_trocas = """
            SELECT 
                COUNT(t.id_troca) as total_trocas,
                COALESCE(SUM(t.valor_total_retorno), 0) as valor_retorno,
                COALESCE(SUM(t.valor_total_substituicao), 0) as valor_substituicao
            FROM trocas t
            WHERE 1=1
        """
        params_trocas = []
        if filtros.get('dataInicio'):
            sql_trocas += " AND DATE(t.data_troca) >= %s"
            params_trocas.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_trocas += " AND DATE(t.data_troca) <= %s"
            params_trocas.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_trocas, params_trocas)
                trocas_metrics = cursor.fetchone()
            total_trocas, valor_trocas_ret, valor_trocas_sub = trocas_metrics if trocas_metrics else (0, 0, 0)
        except:
            total_trocas, valor_trocas_ret, valor_trocas_sub = 0, 0, 0
        
        # ===== INDICADORES DE COMANDAS =====
        sql_comandas = """
            SELECT 
                COUNT(c.id) as total_comandas,
                COALESCE(SUM(c.total), 0) as valor_total,
                COALESCE(SUM(c.desconto), 0) as total_desconto
            FROM comandas c
            WHERE 1=1
        """
        params_comandas = []
        if filtros.get('dataInicio'):
            sql_comandas += " AND DATE(c.data_abertura) >= %s"
            params_comandas.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_comandas += " AND DATE(c.data_abertura) <= %s"
            params_comandas.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_comandas, params_comandas)
                comandas_metrics = cursor.fetchone()
            total_comandas, valor_comandas, desconto_comandas = comandas_metrics if comandas_metrics else (0, 0, 0)
        except:
            total_comandas, valor_comandas, desconto_comandas = 0, 0, 0
        
        # ===== INDICADORES DE DEVOLUÇÕES =====
        sql_devolucoes = """
            SELECT 
                COUNT(d.id_devolucao) as total_devolucoes,
                COALESCE(SUM(d.valor_total_devolucao), 0) as valor_total
            FROM devolucoes d
            WHERE 1=1
        """
        params_devolucoes = []
        
        if filtros.get('dataInicio'):
            sql_devolucoes += " AND DATE(d.data_devolucao) >= %s"
            params_devolucoes.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_devolucoes += " AND DATE(d.data_devolucao) <= %s"
            params_devolucoes.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_devolucoes, params_devolucoes)
                devolucoes_metrics = cursor.fetchone()
            total_devolucoes, valor_devolucoes = devolucoes_metrics if devolucoes_metrics else (0, 0)
        except:
            total_devolucoes, valor_devolucoes = 0, 0
        
        # ===== INDICADORES DE TROCAS =====
        sql_trocas = """
            SELECT 
                COUNT(t.id_troca) as total_trocas,
                COALESCE(SUM(t.valor_total_retorno), 0) as valor_retorno,
                COALESCE(SUM(t.valor_total_substituicao), 0) as valor_substituicao
            FROM trocas t
            WHERE 1=1
        """
        params_trocas = []
        
        if filtros.get('dataInicio'):
            sql_trocas += " AND DATE(t.data_troca) >= %s"
            params_trocas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_trocas += " AND DATE(t.data_troca) <= %s"
            params_trocas.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_trocas, params_trocas)
                trocas_metrics = cursor.fetchone()
            total_trocas, valor_trocas_ret, valor_trocas_sub = trocas_metrics if trocas_metrics else (0, 0, 0)
        except:
            total_trocas, valor_trocas_ret, valor_trocas_sub = 0, 0, 0
        
        # ===== INDICADORES DE COMANDAS =====
        sql_comandas = """
            SELECT 
                COUNT(c.id) as total_comandas,
                COALESCE(SUM(c.total), 0) as valor_total,
                COALESCE(SUM(c.desconto), 0) as total_desconto
            FROM comandas c
            WHERE 1=1
        """
        params_comandas = []
        
        if filtros.get('dataInicio'):
            sql_comandas += " AND DATE(c.data_abertura) >= %s"
            params_comandas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_comandas += " AND DATE(c.data_abertura) <= %s"
            params_comandas.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_comandas, params_comandas)
                comandas_metrics = cursor.fetchone()
            total_comandas, valor_comandas, desconto_comandas = comandas_metrics if comandas_metrics else (0, 0, 0)
        except:
            total_comandas, valor_comandas, desconto_comandas = 0, 0, 0
        
        # ===== RENDERIZAR INDICADORES PRINCIPAIS =====
        titulo_style = ParagraphStyle(
            'TituloSecao',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#388e3c'),
            spaceAfter=10,
            spaceBefore=15
        )
        
        indicador_style = ParagraphStyle(
            'Indicador',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=6
        )
        
        # Seção: Indicadores Gerais
        elements.append(Paragraph('<b>📊 INDICADORES GERAIS</b>', titulo_style))
        
        # Calcular margem e lucratividade
        margem = float(valor_vendas or 0) - float(valor_compras or 0)
        margem_cor = '#2e7d32' if margem >= 0 else '#d32f2f'
        
        # Tabela de indicadores gerais
        data_indicadores = [
            ['Indicador', 'Valor'],
            ['💰 Faturamento Total (Vendas)', f"R$ {float(valor_vendas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['💸 Gastos com Compras', f"R$ {float(valor_compras or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['💵 Margem (Faturamento - Compras)', f"R$ {margem:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['🛒 Total de Vendas', str(total_vendas or 0)],
            ['📦 Total de Compras', str(total_compras or 0)],
            ['↩️ Total de Devoluções', str(total_devolucoes or 0)],
            ['🔄 Total de Trocas', str(total_trocas or 0)],
            ['📋 Total de Comandas', str(total_comandas or 0)],
            ['👥 Clientes Ativos', str(clientes_ativos or 0)],
            ['🏭 Fornecedores Ativos', str(fornecedores_ativos or 0)],
            ['🎯 Ticket Médio', f"R$ {float(ticket_medio or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['💝 Valor de Devoluções', f"R$ {float(valor_devolucoes or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['💳 Faturamento de Comandas', f"R$ {float(valor_comandas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
        ]
        
        table_ind = Table(data_indicadores, colWidths=[9*cm, 8*cm])
        table_ind.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#388e3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table_ind)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== TOP 5 PRODUTOS =====
        elements.append(Paragraph('<b>🏆 TOP 5 PRODUTOS MAIS VENDIDOS</b>', titulo_style))
        
        if top_produtos:
            data_produtos = [['#', 'Produto', 'Qtd Vendida', 'Valor Total']]
            for idx, (nome, qtd, valor) in enumerate(top_produtos, 1):
                nome_fmt = nome[:50] + '...' if len(nome) > 50 else nome
                qtd_fmt = f"{float(qtd):,.1f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                valor_fmt = f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                data_produtos.append([str(idx), nome_fmt, qtd_fmt, valor_fmt])
            
            table_prod = Table(data_produtos, colWidths=[1*cm, 9*cm, 3*cm, 4*cm])
            table_prod.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#388e3c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(table_prod)
        else:
            elements.append(Paragraph('<i>Nenhum produto vendido no período.</i>', styles['Normal']))
        
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== TOP 5 CLIENTES =====
        elements.append(Paragraph('<b>👥 TOP 5 CLIENTES</b>', titulo_style))
        
        if top_clientes:
            data_clientes = [['#', 'Cliente', 'Vendas', 'Valor Total']]
            for idx, (nome, vendas, valor) in enumerate(top_clientes, 1):
                nome_fmt = nome[:50] + '...' if len(nome) > 50 else nome
                data_clientes.append([
                    str(idx), 
                    nome_fmt, 
                    str(vendas),
                    f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                ])
            
            table_cli = Table(data_clientes, colWidths=[1*cm, 9*cm, 3*cm, 4*cm])
            table_cli.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#388e3c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (3, -1), 'RIGHT'),
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(table_cli)
        else:
            elements.append(Paragraph('<i>Nenhum cliente encontrado no período.</i>', styles['Normal']))
        
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: DEVOLUÇÕES =====
        elements.append(Paragraph('<b>↩️ DEVOLUÇÕES</b>', titulo_style))
        
        if total_devolucoes > 0:
            data_devol = [
                ['Indicador', 'Valor'],
                ['Total de Devoluções', str(int(total_devolucoes or 0))],
                ['Valor Total Devolvido', f"R$ {float(valor_devolucoes or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
                ['Ticket Médio (Devolução)', f"R$ {float(valor_devolucoes or 1) / float(total_devolucoes or 1):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
            ]
            
            table_devol = Table(data_devol, colWidths=[9*cm, 8*cm])
            table_devol.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffebee')]),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table_devol)
        else:
            elements.append(Paragraph('<i>Nenhuma devolução registrada no período.</i>', styles['Normal']))
        
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: TROCAS =====
        elements.append(Paragraph('<b>🔄 TROCAS</b>', titulo_style))
        
        if total_trocas > 0:
            valor_liquido_trocas = float(valor_trocas_sub or 0) - float(valor_trocas_ret or 0)
            data_trocas = [
                ['Indicador', 'Valor'],
                ['Total de Trocas', str(int(total_trocas or 0))],
                ['Valor Total Retornado', f"R$ {float(valor_trocas_ret or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
                ['Valor Total Substituído', f"R$ {float(valor_trocas_sub or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
                ['Diferença Líquida', f"R$ {valor_liquido_trocas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
            ]
            
            table_trocas = Table(data_trocas, colWidths=[9*cm, 8*cm])
            table_trocas.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6a1b9a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3e5f5')]),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table_trocas)
        else:
            elements.append(Paragraph('<i>Nenhuma troca registrada no período.</i>', styles['Normal']))
        
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: COMANDAS =====
        elements.append(Paragraph('<b>📋 COMANDAS</b>', titulo_style))
        
        if total_comandas > 0:
            data_comandas = [
                ['Indicador', 'Valor'],
                ['Total de Comandas', str(int(total_comandas or 0))],
                ['Faturamento Total', f"R$ {float(valor_comandas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
                ['Desconto Total Concedido', f"R$ {float(desconto_comandas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
                ['Ticket Médio (Comanda)', f"R$ {float(valor_comandas or 1) / float(total_comandas or 1):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
            ]
            
            table_comandas = Table(data_comandas, colWidths=[9*cm, 8*cm])
            table_comandas.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00796b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e0f2f1')]),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table_comandas)
        else:
            elements.append(Paragraph('<i>Nenhuma comanda registrada no período.</i>', styles['Normal']))
        
        # Mensagem final
        if not top_produtos and not top_clientes:
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph(
                '<i>Selecione um período com dados para visualizar a análise de desempenho.</i>',
                styles['Normal']
            ))
        
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_desempenho_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_consolidado(self, filtros):
        """Gera relatório consolidado com visão geral de todas as operações"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório Consolidado',
            'Visão Geral Completa das Operações do Sistema'
        )
        
        # Mostrar período de análise
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Período:</b> {' até '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Estilos personalizados
        titulo_style = ParagraphStyle(
            'TituloSecao',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#7b1fa2'),
            spaceAfter=10,
            spaceBefore=15
        )
        
        # ===== VENDAS =====
        sql_vendas = """
            SELECT 
                COUNT(v.id_venda) as total_vendas,
                COALESCE(SUM((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as valor_total,
                COUNT(DISTINCT v.id_cliente) as clientes_distintos
            FROM vendas v
            WHERE 1=1
        """
        params_vendas = []
        
        if filtros.get('dataInicio'):
            sql_vendas += " AND DATE(v.data_documento) >= %s"
            params_vendas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_vendas += " AND DATE(v.data_documento) <= %s"
            params_vendas.append(filtros['dataFim'])
        
        with connection.cursor() as cursor:
            cursor.execute(sql_vendas, params_vendas)
            vendas_data = cursor.fetchone()
        
        total_vendas, valor_vendas, clientes_vendas = vendas_data
        
        # ===== COMPRAS =====
        sql_compras = """
            SELECT 
                COUNT(c.id_compra) as total_compras,
                COALESCE(SUM(c.valor_total_nota), 0) as valor_total,
                COUNT(DISTINCT c.id_fornecedor) as fornecedores_distintos
            FROM compras c
            WHERE 1=1
        """
        params_compras = []
        
        if filtros.get('dataInicio'):
            sql_compras += " AND DATE(c.data_movimento_entrada) >= %s"
            params_compras.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_compras += " AND DATE(c.data_movimento_entrada) <= %s"
            params_compras.append(filtros['dataFim'])
        
        with connection.cursor() as cursor:
            cursor.execute(sql_compras, params_compras)
            compras_data = cursor.fetchone()
        
        total_compras, valor_compras, fornecedores_compras = compras_data
        
        # ===== ESTOQUE =====
        sql_estoque = """
            SELECT 
                COUNT(DISTINCT p.id_produto) as total_produtos,
                COALESCE(SUM(e.quantidade), 0) as quantidade_total,
                COALESCE(SUM(e.quantidade * e.valor_venda), 0) as valor_estoque
            FROM produtos p
            LEFT JOIN estoque e ON p.id_produto = e.id_produto
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql_estoque)
            estoque_data = cursor.fetchone()
        
        total_produtos, qtd_estoque, valor_estoque = estoque_data
        
        # ===== CLIENTES =====
        sql_clientes = """
            SELECT COUNT(*) as total_clientes
            FROM clientes
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql_clientes)
            total_clientes = cursor.fetchone()[0]
        
        # ===== FORNECEDORES =====
        sql_fornecedores = """
            SELECT COUNT(*) as total_fornecedores
            FROM fornecedores
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql_fornecedores)
            total_fornecedores = cursor.fetchone()[0]
        
        # ===== DEVOLUÇÕES =====
        sql_devolucoes = """
            SELECT 
                COUNT(d.id_devolucao) as total_devolucoes,
                COALESCE(SUM(d.valor_total_devolucao), 0) as valor_total
            FROM devolucoes d
            WHERE 1=1
        """
        params_devolucoes = []
        
        if filtros.get('dataInicio'):
            sql_devolucoes += " AND DATE(d.data_devolucao) >= %s"
            params_devolucoes.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_devolucoes += " AND DATE(d.data_devolucao) <= %s"
            params_devolucoes.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_devolucoes, params_devolucoes)
                devolucoes_metrics = cursor.fetchone()
            total_devolucoes, valor_devolucoes = devolucoes_metrics if devolucoes_metrics else (0, 0)
        except:
            total_devolucoes, valor_devolucoes = 0, 0
        
        # ===== TROCAS =====
        sql_trocas = """
            SELECT 
                COUNT(t.id_troca) as total_trocas,
                COALESCE(SUM(t.valor_total_retorno), 0) as valor_retorno,
                COALESCE(SUM(t.valor_total_substituicao), 0) as valor_substituicao
            FROM trocas t
            WHERE 1=1
        """
        params_trocas = []
        
        if filtros.get('dataInicio'):
            sql_trocas += " AND DATE(t.data_troca) >= %s"
            params_trocas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_trocas += " AND DATE(t.data_troca) <= %s"
            params_trocas.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_trocas, params_trocas)
                trocas_metrics = cursor.fetchone()
            total_trocas, valor_trocas_ret, valor_trocas_sub = trocas_metrics if trocas_metrics else (0, 0, 0)
        except:
            total_trocas, valor_trocas_ret, valor_trocas_sub = 0, 0, 0
        
        # ===== COMANDAS =====
        sql_comandas = """
            SELECT 
                COUNT(c.id) as total_comandas,
                COALESCE(SUM(c.total), 0) as valor_total,
                COALESCE(SUM(c.desconto), 0) as total_desconto
            FROM comandas c
            WHERE 1=1
        """
        params_comandas = []
        
        if filtros.get('dataInicio'):
            sql_comandas += " AND DATE(c.data_abertura) >= %s"
            params_comandas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_comandas += " AND DATE(c.data_abertura) <= %s"
            params_comandas.append(filtros['dataFim'])
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_comandas, params_comandas)
                comandas_metrics = cursor.fetchone()
            total_comandas, valor_comandas, desconto_comandas = comandas_metrics if comandas_metrics else (0, 0, 0)
        except:
            total_comandas, valor_comandas, desconto_comandas = 0, 0, 0
        
        # ===== SEÇÃO: RESUMO FINANCEIRO =====
        elements.append(Paragraph('<b>💰 RESUMO FINANCEIRO</b>', titulo_style))
        
        margem_bruta = float(valor_vendas or 0) - float(valor_compras or 0)
        margem_cor = '#2e7d32' if margem_bruta >= 0 else '#d32f2f'
        
        data_financeiro = [
            ['Indicador', 'Valor'],
            ['Faturamento (Vendas)', f"R$ {float(valor_vendas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Investimento (Compras)', f"R$ {float(valor_compras or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Margem Bruta', f"R$ {margem_bruta:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Valor em Estoque', f"R$ {float(valor_estoque or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
        ]
        
        table_fin = Table(data_financeiro, colWidths=[9*cm, 8*cm])
        table_fin.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7b1fa2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table_fin)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: OPERAÇÕES =====
        elements.append(Paragraph('<b>📊 OPERAÇÕES</b>', titulo_style))
        
        data_operacoes = [
            ['Operação', 'Quantidade', 'Valor Total'],
            ['🛒 Vendas Realizadas', str(total_vendas or 0), f"R$ {float(valor_vendas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['📦 Compras Realizadas', str(total_compras or 0), f"R$ {float(valor_compras or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['↩️ Devoluções Registradas', str(total_devolucoes or 0), f"R$ {float(valor_devolucoes or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['🔄 Trocas Realizadas', str(total_trocas or 0), f"R$ {float(valor_trocas_sub or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['📋 Comandas Fechadas', str(total_comandas or 0), f"R$ {float(valor_comandas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ]
        
        table_op = Table(data_operacoes, colWidths=[7*cm, 5*cm, 5*cm])
        table_op.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7b1fa2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table_op)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: CADASTROS =====
        elements.append(Paragraph('<b>📋 CADASTROS</b>', titulo_style))
        
        data_cadastros = [
            ['Tipo', 'Total Cadastrado', 'Ativos no Período'],
            ['👥 Clientes', str(total_clientes or 0), str(clientes_vendas or 0)],
            ['🏭 Fornecedores', str(total_fornecedores or 0), str(fornecedores_compras or 0)],
            ['📦 Produtos', str(total_produtos or 0), '-'],
        ]
        
        table_cad = Table(data_cadastros, colWidths=[7*cm, 5*cm, 5*cm])
        table_cad.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7b1fa2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (2, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table_cad)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: ESTOQUE =====
        elements.append(Paragraph('<b>📦 ESTOQUE</b>', titulo_style))
        
        data_estoque_info = [
            ['Indicador', 'Valor'],
            ['Total de Produtos', str(total_produtos or 0)],
            ['Quantidade Total em Estoque', f"{float(qtd_estoque or 0):,.1f} Un".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Valor Total do Estoque', f"R$ {float(valor_estoque or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ]
        
        table_est = Table(data_estoque_info, colWidths=[9*cm, 8*cm])
        table_est.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7b1fa2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table_est)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: DEVOLUÇÕES =====
        elements.append(Paragraph('<b>↩️ DEVOLUÇÕES</b>', titulo_style))
        
        data_devol = [
            ['Indicador', 'Valor'],
            ['Total de Devoluções', str(int(total_devolucoes or 0))],
            ['Valor Total Devolvido', f"R$ {float(valor_devolucoes or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ]
        
        if total_devolucoes > 0:
            ticket_devolucao = float(valor_devolucoes or 0) / float(total_devolucoes)
            data_devol.append(['Ticket Médio (Devolução)', f"R$ {ticket_devolucao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        table_devol = Table(data_devol, colWidths=[9*cm, 8*cm])
        table_devol.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffebee')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table_devol)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: TROCAS =====
        elements.append(Paragraph('<b>🔄 TROCAS</b>', titulo_style))
        
        valor_liquido_trocas = float(valor_trocas_sub or 0) - float(valor_trocas_ret or 0)
        data_trocas = [
            ['Indicador', 'Valor'],
            ['Total de Trocas', str(int(total_trocas or 0))],
            ['Valor Total Retornado', f"R$ {float(valor_trocas_ret or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Valor Total Substituído', f"R$ {float(valor_trocas_sub or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Diferença Líquida', f"R$ {valor_liquido_trocas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ]
        
        table_trocas = Table(data_trocas, colWidths=[9*cm, 8*cm])
        table_trocas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6a1b9a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3e5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table_trocas)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: COMANDAS =====
        elements.append(Paragraph('<b>📋 COMANDAS</b>', titulo_style))
        
        data_comandas = [
            ['Indicador', 'Valor'],
            ['Total de Comandas', str(int(total_comandas or 0))],
            ['Faturamento Total', f"R$ {float(valor_comandas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
            ['Desconto Total Concedido', f"R$ {float(desconto_comandas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ]
        
        if total_comandas > 0:
            ticket_comanda = float(valor_comandas or 0) / float(total_comandas)
            data_comandas.append(['Ticket Médio (Comanda)', f"R$ {ticket_comanda:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        table_comandas = Table(data_comandas, colWidths=[9*cm, 8*cm])
        table_comandas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00796b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e0f2f1')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table_comandas)
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== RESUMO EXECUTIVO =====
        elements.append(Paragraph('<b>📈 RESUMO EXECUTIVO</b>', titulo_style))
        
        ticket_medio = float(valor_vendas or 0) / float(total_vendas or 1) if total_vendas else 0
        
        resumo_style = ParagraphStyle(
            'ResumoExecutivo',
            parent=styles['Normal'],
            fontSize=9,
            leading=13,
            spaceAfter=5
        )
        
        elements.append(Paragraph(
            f'• <b>Ticket Médio de Venda:</b> R$ {ticket_medio:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        
        if total_produtos > 0:
            valor_medio_produto = float(valor_estoque or 0) / float(total_produtos)
            elements.append(Paragraph(
                f'• <b>Valor Médio por Produto em Estoque:</b> R$ {valor_medio_produto:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                resumo_style
            ))
        
        if total_clientes > 0:
            taxa_conversao = (float(clientes_vendas or 0) / float(total_clientes)) * 100
            elements.append(Paragraph(
                f'• <b>Taxa de Clientes Ativos:</b> {taxa_conversao:.1f}%'.replace('.', ','),
                resumo_style
            ))
        
        if margem_bruta >= 0:
            elements.append(Paragraph(
                f'• <b>Situação Financeira:</b> <font color="#2e7d32">Positiva (Margem de R$ {margem_bruta:,.2f})</font>'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                resumo_style
            ))
        else:
            elements.append(Paragraph(
                f'• <b>Situação Financeira:</b> <font color="#d32f2f">Atenção (Déficit de R$ {abs(margem_bruta):,.2f})</font>'.replace(',', 'X').replace('.', ',').replace('X', '.'),
                resumo_style
            ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_consolidado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_comissoes(self, filtros):
        """Gera relatório de comissões por vendedor"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Comissões por Vendedor',
            'Comissões de Vendas e Contas Recebidas'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('vendedor'):
            filtros_text.append(f"Vendedor ID: {filtros['vendedor']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Estilos personalizados
        titulo_style = ParagraphStyle(
            'TituloSecao',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#f57c00'),
            spaceAfter=10,
            spaceBefore=15
        )
        
        # ===== COMISSÕES POR VENDEDOR =====
        # Query corrigida para usar id_vendedor1 e mostrar detalhes das vendas
        sql_comissoes = """
            SELECT 
                v.id_vendedor,
                v.nome as vendedor_nome,
                v.percentual_comissao,
                
                -- Comissões de Vendas (baseado no valor total da venda)
                COUNT(DISTINCT vd.id_venda) as total_vendas,
                COALESCE(SUM(vd.valor_total), 0) as valor_total_vendas,
                COALESCE(SUM(vd.valor_total * (v.percentual_comissao / 100)), 0) as comissao_vendas,
                
                -- Comissões de Contas Recebidas (baseado no valor pago)
                COUNT(DISTINCT CASE WHEN fc.status_conta = 'Paga' THEN fc.id_conta END) as total_contas_recebidas,
                COALESCE(SUM(
                    CASE WHEN fc.status_conta = 'Paga' 
                    THEN fc.valor_liquidado 
                    ELSE 0 END
                ), 0) as valor_recebido,
                COALESCE(SUM(
                    CASE WHEN fc.status_conta = 'Paga' 
                    THEN fc.valor_liquidado * (v.percentual_comissao / 100)
                    ELSE 0 END
                ), 0) as comissao_recebimentos
                
            FROM vendedores v
            LEFT JOIN vendas vd ON v.id_vendedor = vd.id_vendedor1
            LEFT JOIN financeiro_contas fc ON vd.id_venda = fc.id_venda_origem 
                AND fc.tipo_conta = 'Receber'
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql_comissoes += " AND DATE(vd.data_documento) >= %s"
            params.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_comissoes += " AND DATE(vd.data_documento) <= %s"
            params.append(filtros['dataFim'])
        
        if filtros.get('vendedor'):
            sql_comissoes += " AND v.id_vendedor = %s"
            params.append(filtros['vendedor'])
        
        sql_comissoes += """
            GROUP BY v.id_vendedor, v.nome, v.percentual_comissao
            HAVING total_vendas > 0 OR total_contas_recebidas > 0
            ORDER BY (comissao_vendas + comissao_recebimentos) DESC
        """
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql_comissoes, params)
                vendedores_comissoes = cursor.fetchall()
        except Exception as e:
            print(f"Erro na query de comissões: {e}")
            raise
        
        # ===== SEÇÃO: COMISSÕES DE VENDAS =====
        elements.append(Paragraph('<b>💰 COMISSÕES DE VENDAS</b>', titulo_style))
        
        data_vendas = [['Vendedor', '% Com.', 'Vendas', 'Valor Vendido', 'Comissão']]
        col_widths_vendas = [5*cm, 1.5*cm, 1.5*cm, 3*cm, 3*cm]
        
        total_vendas_geral = 0
        total_valor_vendido_geral = 0
        total_comissao_vendas_geral = 0
        
        for vend in vendedores_comissoes:
            id_vend, nome, perc_comissao, qtd_vendas, valor_vendas, comissao_vendas, qtd_recebidas, valor_recebido, comissao_receb = vend
            
            if qtd_vendas > 0:
                total_vendas_geral += int(qtd_vendas or 0)
                total_valor_vendido_geral += float(valor_vendas or 0)
                total_comissao_vendas_geral += float(comissao_vendas or 0)
                
                nome_fmt = nome[:35] + '...' if len(nome) > 35 else nome
                perc_fmt = f"{float(perc_comissao or 0):.1f}%"
                valor_fmt = f"R$ {float(valor_vendas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                comissao_fmt = f"R$ {float(comissao_vendas or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                data_vendas.append([
                    nome_fmt,
                    perc_fmt,
                    str(int(qtd_vendas)),
                    valor_fmt,
                    comissao_fmt
                ])
        
        if len(data_vendas) == 1:
            data_vendas.append(['-', '-', '-', 'Nenhuma venda encontrada', '-'])
        
        table_vendas = Table(data_vendas, colWidths=col_widths_vendas)
        table_vendas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f57c00')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table_vendas)
        elements.append(Spacer(1, 0.5*cm))
        
        # Resumo Vendas
        resumo_style = ParagraphStyle(
            'Resumo',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#f57c00'),
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph(
            f'Total de Vendas: {total_vendas_geral} | ' +
            f'Valor Total Vendido: R$ {total_valor_vendido_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.') +
            f' | Comissão Total: R$ {total_comissao_vendas_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        
        elements.append(Spacer(1, 0.8*cm))
        
        # ===== SEÇÃO: DETALHAMENTO DE VENDAS POR VENDEDOR =====
        elements.append(Paragraph('<b>📋 DETALHAMENTO DAS VENDAS</b>', titulo_style))
        
        # Query para buscar vendas detalhadas
        sql_vendas_detalhadas = """
            SELECT 
                v.nome as vendedor_nome,
                v.percentual_comissao,
                vd.id_venda,
                vd.numero_documento,
                DATE_FORMAT(vd.data_documento, '%d/%m/%Y %H:%i') as data_venda,
                c.nome_razao_social as cliente_nome,
                vd.valor_total,
                (vd.valor_total * (v.percentual_comissao / 100)) as comissao_venda
            FROM vendedores v
            INNER JOIN vendas vd ON v.id_vendedor = vd.id_vendedor1
            LEFT JOIN clientes c ON vd.id_cliente = c.id_cliente
            WHERE 1=1
        """
        
        params_detalhadas = []
        
        if filtros.get('dataInicio'):
            sql_vendas_detalhadas += " AND DATE(vd.data_documento) >= %s"
            params_detalhadas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_vendas_detalhadas += " AND DATE(vd.data_documento) <= %s"
            params_detalhadas.append(filtros['dataFim'])
        
        if filtros.get('vendedor'):
            sql_vendas_detalhadas += " AND v.id_vendedor = %s"
            params_detalhadas.append(filtros['vendedor'])
        
        sql_vendas_detalhadas += " ORDER BY v.nome, vd.data_documento DESC"
        
        with connection.cursor() as cursor:
            cursor.execute(sql_vendas_detalhadas, params_detalhadas)
            vendas_detalhadas = cursor.fetchall()
        
        if vendas_detalhadas:
            # Agrupar vendas por vendedor
            vendas_por_vendedor = {}
            for venda in vendas_detalhadas:
                vendedor_nome = venda[0]
                if vendedor_nome not in vendas_por_vendedor:
                    vendas_por_vendedor[vendedor_nome] = []
                vendas_por_vendedor[vendedor_nome].append(venda)
            
            # Criar tabela para cada vendedor
            for vendedor_nome, vendas in vendas_por_vendedor.items():
                perc_comissao = vendas[0][1] if vendas else 0
                
                # Título do vendedor
                vendedor_title_style = ParagraphStyle(
                    'VendedorTitle',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#1976d2'),
                    spaceAfter=8,
                    spaceBefore=10,
                    fontName='Helvetica-Bold'
                )
                elements.append(Paragraph(
                    f'Vendedor: {vendedor_nome} (Comissão: {float(perc_comissao):.1f}%)',
                    vendedor_title_style
                ))
                
                # Tabela de vendas do vendedor
                data_vendas_det = [['Nº Doc', 'Data', 'Cliente', 'Valor Venda', 'Comissão']]
                col_widths_det = [2*cm, 2.5*cm, 5*cm, 2.5*cm, 2.5*cm]
                
                total_vendedor = 0
                total_comissao_vendedor = 0
                
                for venda in vendas:
                    _, _, id_venda, num_doc, data_venda, cliente_nome, valor_total, comissao = venda
                    
                    total_vendedor += float(valor_total or 0)
                    total_comissao_vendedor += float(comissao or 0)
                    
                    num_doc_fmt = num_doc or f'#{id_venda}'
                    cliente_fmt = (cliente_nome[:30] + '...') if cliente_nome and len(cliente_nome) > 30 else (cliente_nome or '-')
                    valor_fmt = f"R$ {float(valor_total or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    comissao_fmt = f"R$ {float(comissao or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    
                    data_vendas_det.append([
                        num_doc_fmt,
                        data_venda,
                        cliente_fmt,
                        valor_fmt,
                        comissao_fmt
                    ])
                
                # Linha de total do vendedor
                valor_total_fmt = f"R$ {total_vendedor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                comissao_total_fmt = f"R$ {total_comissao_vendedor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                data_vendas_det.append([
                    '', '', 'TOTAL:', valor_total_fmt, comissao_total_fmt
                ])
                
                table_vendas_det = Table(data_vendas_det, colWidths=col_widths_det)
                table_vendas_det.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ALIGN', (0, 1), (1, -1), 'CENTER'),
                    ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                    ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 1), (-1, -2), 5),
                    ('BOTTOMPADDING', (0, 1), (-1, -2), 5),
                    # Linha de total
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd')),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, -1), (-1, -1), 8),
                    ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
                    ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#1976d2')),
                    ('TOPPADDING', (0, -1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
                ]))
                
                elements.append(table_vendas_det)
                elements.append(Spacer(1, 0.3*cm))
        else:
            elements.append(Paragraph('Nenhuma venda encontrada no período.', styles['Normal']))
        
        elements.append(Spacer(1, 1*cm))
        
        # ===== SEÇÃO: COMISSÕES DE CONTAS RECEBIDAS =====
        elements.append(Paragraph('<b>💵 COMISSÕES DE CONTAS RECEBIDAS</b>', titulo_style))
        
        data_receb = [['Vendedor', '% Com.', 'Contas', 'Valor Recebido', 'Comissão']]
        col_widths_receb = [5*cm, 1.5*cm, 1.5*cm, 3*cm, 3*cm]
        
        total_contas_geral = 0
        total_valor_recebido_geral = 0
        total_comissao_receb_geral = 0
        
        for vend in vendedores_comissoes:
            id_vend, nome, perc_comissao, qtd_vendas, valor_vendas, comissao_vendas, qtd_recebidas, valor_recebido, comissao_receb = vend
            
            if qtd_recebidas > 0:
                total_contas_geral += int(qtd_recebidas or 0)
                total_valor_recebido_geral += float(valor_recebido or 0)
                total_comissao_receb_geral += float(comissao_receb or 0)
                
                nome_fmt = nome[:35] + '...' if len(nome) > 35 else nome
                perc_fmt = f"{float(perc_comissao or 0):.1f}%"
                valor_fmt = f"R$ {float(valor_recebido or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                comissao_fmt = f"R$ {float(comissao_receb or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                
                data_receb.append([
                    nome_fmt,
                    perc_fmt,
                    str(int(qtd_recebidas)),
                    valor_fmt,
                    comissao_fmt
                ])
        
        if len(data_receb) == 1:
            data_receb.append(['-', '-', '-', 'Nenhuma conta recebida encontrada', '-'])
        
        table_receb = Table(data_receb, colWidths=col_widths_receb)
        table_receb.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#388e3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table_receb)
        elements.append(Spacer(1, 0.5*cm))
        
        # Resumo Recebimentos
        elements.append(Paragraph(
            f'Total de Contas: {total_contas_geral} | ' +
            f'Valor Total Recebido: R$ {total_valor_recebido_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.') +
            f' | Comissão Total: R$ {total_comissao_receb_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.'),
            resumo_style
        ))
        
        elements.append(Spacer(1, 1*cm))
        
        # ===== RESUMO TOTAL DE COMISSÕES =====
        elements.append(Paragraph('<b>📊 RESUMO TOTAL</b>', titulo_style))
        
        total_comissoes_geral = total_comissao_vendas_geral + total_comissao_receb_geral
        
        data_resumo = [
            ['Tipo', 'Quantidade', 'Valor Base', 'Comissão'],
            [
                'Vendas',
                str(total_vendas_geral),
                f"R$ {total_valor_vendido_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {total_comissao_vendas_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ],
            [
                'Contas Recebidas',
                str(total_contas_geral),
                f"R$ {total_valor_recebido_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {total_comissao_receb_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ],
            [
                'TOTAL GERAL',
                '',
                '',
                f"R$ {total_comissoes_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ]
        ]
        
        table_resumo = Table(data_resumo, colWidths=[4*cm, 3*cm, 4*cm, 3*cm])
        table_resumo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f57c00')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f57c00')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table_resumo)
        
        if not vendedores_comissoes:
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph(
                '<i>Nenhuma comissão encontrada para o período selecionado.</i>',
                styles['Normal']
            ))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_comissoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    # ===== MÉTODOS PARA GERAÇÃO DE EXCEL =====
    
    def criar_excel_base(self, titulo, colunas, dados, nome_planilha='Dados'):
        """Cria uma planilha Excel formatada"""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Biblioteca openpyxl não instalada. Execute: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nome_planilha
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells(f'A1:{chr(64 + len(colunas))}1')
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Cabeçalhos
        for idx, col in enumerate(colunas, 1):
            cell = ws.cell(row=2, column=idx)
            cell.value = col
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados
        for row_idx, row_data in enumerate(dados, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Pular MergedCell (células mescladas não têm column_letter)
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def gerar_relatorio_vendas_excel(self, filtros):
        """Gera relatório de vendas em Excel"""
        from django.db import connection
        
        # Buscar vendas
        sql = """
            SELECT 
                v.numero_documento,
                v.data_documento,
                c.nome_razao_social,
                COALESCE(
                    (SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda),
                    0
                ) as valor,
                v.status_venda
            FROM vendas v
            LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(v.data_documento) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(v.data_documento) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('cliente'):
            sql += " AND v.id_cliente = %s"
            params.append(filtros['cliente'])
        
        sql += " ORDER BY v.data_documento DESC LIMIT 1000"
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            vendas = cursor.fetchall()
        
        # Preparar dados
        dados = []
        total = 0
        for venda in vendas:
            numero_doc, data_venda, cliente, valor, status_venda = venda
            data_fmt = data_venda.strftime('%d/%m/%Y') if data_venda else '-'
            valor_num = float(valor or 0)
            total += valor_num
            dados.append([
                numero_doc or '-',
                data_fmt,
                cliente or 'N/A',
                valor_num,
                status_venda or 'Pendente'
            ])
        
        # Adicionar linha de total
        dados.append(['', '', 'TOTAL', total, ''])
        
        wb = self.criar_excel_base(
            'Relatório de Vendas',
            ['Documento', 'Data', 'Cliente', 'Valor', 'Status'],
            dados,
            'Vendas'
        )
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_vendas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_estoque_excel(self, filtros):
        """Gera relatório de estoque em Excel"""
        from django.db import connection
        
        # Buscar estoque
        sql = """
            SELECT 
                p.codigo_produto,
                p.nome_produto,
                COALESCE(e.id_deposito, 0) as deposito,
                COALESCE(e.quantidade, 0) as quantidade,
                COALESCE(e.valor_venda, 0) as valor_venda,
                COALESCE(e.valor_ultima_compra, 0) as valor_compra
            FROM produtos p
            LEFT JOIN estoque e ON p.id_produto = e.id_produto
            WHERE 1=1
        """
        params = []
        
        if filtros.get('produto'):
            sql += " AND p.id_produto = %s"
            params.append(filtros['produto'])
        if filtros.get('deposito'):
            sql += " AND e.id_deposito = %s"
            params.append(filtros['deposito'])
        
        sql += " ORDER BY p.nome_produto"
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            estoque = cursor.fetchall()
        
        # Preparar dados
        dados = []
        for item in estoque:
            codigo, nome, deposito, qtd, vl_venda, vl_compra = item
            if float(qtd) > 0:  # Apenas itens com estoque
                dados.append([
                    codigo or '-',
                    nome or '-',
                    f'Depósito {deposito}',
                    float(qtd),
                    float(vl_venda),
                    float(vl_compra)
                ])
        
        wb = self.criar_excel_base(
            'Relatório de Estoque',
            ['Código', 'Produto', 'Depósito', 'Quantidade', 'Vlr Venda', 'Vlr Compra'],
            dados,
            'Estoque'
        )
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_compras_excel(self, filtros):
        """Gera relatório de compras em Excel"""
        return self.gerar_excel_simples('compras', filtros)
    
    def gerar_relatorio_financeiro_excel(self, filtros):
        """Gera relatório financeiro em Excel"""
        return self.gerar_excel_simples('financeiro', filtros)
    
    def gerar_relatorio_clientes_excel(self, filtros):
        """Gera relatório de clientes em Excel"""
        return self.gerar_excel_simples('clientes', filtros)
    
    def gerar_relatorio_produtos_excel(self, filtros):
        """Gera relatório de produtos em Excel"""
        return self.gerar_excel_simples('produtos', filtros)
    
    def gerar_relatorio_desempenho_excel(self, filtros):
        """Gera relatório de desempenho em Excel"""
        from django.db import connection

        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Exportação para Excel requer a biblioteca openpyxl. Execute: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # ===== INDICADORES DE VENDAS =====
        sql_vendas = """
            SELECT 
                COUNT(DISTINCT v.id_venda) as total_vendas,
                COALESCE(SUM((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as valor_total,
                COUNT(DISTINCT v.id_cliente) as clientes_ativos,
                COALESCE(AVG((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as ticket_medio
            FROM vendas v
            WHERE 1=1
        """
        params_vendas = []
        if filtros.get('dataInicio'):
            sql_vendas += " AND DATE(v.data_documento) >= %s"
            params_vendas.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_vendas += " AND DATE(v.data_documento) <= %s"
            params_vendas.append(filtros['dataFim'])

        with connection.cursor() as cursor:
            cursor.execute(sql_vendas, params_vendas)
            vendas_metrics = cursor.fetchone() or (0, 0, 0, 0)

        total_vendas, valor_vendas, clientes_ativos, ticket_medio = vendas_metrics

        # ===== INDICADORES DE COMPRAS =====
        sql_compras = """
            SELECT 
                COUNT(c.id_compra) as total_compras,
                COALESCE(SUM(c.valor_total_nota), 0) as valor_total,
                COUNT(DISTINCT c.id_fornecedor) as fornecedores_ativos
            FROM compras c
            WHERE 1=1
        """
        params_compras = []
        if filtros.get('dataInicio'):
            sql_compras += " AND DATE(c.data_movimento_entrada) >= %s"
            params_compras.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_compras += " AND DATE(c.data_movimento_entrada) <= %s"
            params_compras.append(filtros['dataFim'])

        with connection.cursor() as cursor:
            cursor.execute(sql_compras, params_compras)
            compras_metrics = cursor.fetchone() or (0, 0, 0)

        total_compras, valor_compras, fornecedores_ativos = compras_metrics

        # ===== TOP 5 PRODUTOS MAIS VENDIDOS =====
        sql_top_produtos = """
            SELECT 
                p.nome_produto,
                SUM(vi.quantidade) as qtd_vendida,
                SUM(vi.valor_total) as valor_total
            FROM venda_itens vi
            JOIN produtos p ON vi.id_produto = p.id_produto
            JOIN vendas v ON vi.id_venda = v.id_venda
            WHERE 1=1
        """
        params_top = []
        if filtros.get('dataInicio'):
            sql_top_produtos += " AND DATE(v.data_documento) >= %s"
            params_top.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_top_produtos += " AND DATE(v.data_documento) <= %s"
            params_top.append(filtros['dataFim'])
        sql_top_produtos += " GROUP BY p.id_produto, p.nome_produto ORDER BY valor_total DESC LIMIT 5"

        with connection.cursor() as cursor:
            cursor.execute(sql_top_produtos, params_top)
            top_produtos = cursor.fetchall()

        # ===== TOP 5 CLIENTES =====
        sql_top_clientes = """
            SELECT 
                c.nome_razao_social,
                COUNT(DISTINCT v.id_venda) as total_vendas,
                COALESCE(SUM((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as valor_total
            FROM clientes c
            JOIN vendas v ON c.id_cliente = v.id_cliente
            WHERE 1=1
        """
        params_clientes = []
        if filtros.get('dataInicio'):
            sql_top_clientes += " AND DATE(v.data_documento) >= %s"
            params_clientes.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_top_clientes += " AND DATE(v.data_documento) <= %s"
            params_clientes.append(filtros['dataFim'])
        sql_top_clientes += " GROUP BY c.id_cliente, c.nome_razao_social ORDER BY valor_total DESC LIMIT 5"

        with connection.cursor() as cursor:
            cursor.execute(sql_top_clientes, params_clientes)
            top_clientes = cursor.fetchall()

        # ===== MONTAR ARQUIVO EXCEL =====
        # Indicadores
        margem = float(valor_vendas or 0) - float(valor_compras or 0)
        ticket_medio_val = float(ticket_medio or 0)

        data_indicadores = [
            ['💰 Faturamento Total', float(valor_vendas or 0)],
            ['💸 Gastos com Compras', float(valor_compras or 0)],
            ['💵 Margem (Faturamento - Compras)', float(margem)],
            ['🛒 Total de Vendas', int(total_vendas or 0)],
            ['📦 Total de Compras', int(total_compras or 0)],
            ['↩️ Total de Devoluções', int(total_devolucoes or 0)],
            ['💝 Valor de Devoluções', float(valor_devolucoes or 0)],
            ['🔄 Total de Trocas', int(total_trocas or 0)],
            ['💳 Faturamento de Comandas', float(valor_comandas or 0)],
            ['📋 Total de Comandas', int(total_comandas or 0)],
            ['👥 Clientes Ativos', int(clientes_ativos or 0)],
            ['🏭 Fornecedores Ativos', int(fornecedores_ativos or 0)],
            ['🎯 Ticket Médio', float(ticket_medio_val)]
        ]

        # Usar helper para criar a primeira planilha (Indicadores)
        wb = self.criar_excel_base(
            'Relatório de Desempenho - Indicadores',
            ['Indicador', 'Valor'],
            data_indicadores,
            'Indicadores'
        )

        # Funções utilitárias locais para criar folhas adicionais com estilo similar
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )

        def criar_planilha_simples(wb, nome, titulo, colunas, linhas):
            ws = wb.create_sheet(title=nome)
            # Título
            last_col = chr(64 + len(colunas))
            try:
                ws.merge_cells(f'A1:{last_col}1')
            except Exception:
                pass
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14)
            ws.row_dimensions[1].height = 25

            # Cabeçalhos
            for idx, col in enumerate(colunas, 1):
                cell = ws.cell(row=2, column=idx)
                cell.value = col
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Dados
            for r_idx, row in enumerate(linhas, start=3):
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = val
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')

            # Ajustar largura de colunas
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    # Pular MergedCell (células mescladas não têm column_letter)
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                
                if column_letter:
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            return ws

        # Top Produtos
        linhas_prod = []
        for idx, item in enumerate(top_produtos, 1):
            nome, qtd, valor = item
            linhas_prod.append([idx, nome, float(qtd or 0), float(valor or 0)])

        criar_planilha_simples(
            wb,
            'Top Produtos',
            'Top 5 Produtos Mais Vendidos',
            ['#', 'Produto', 'Qtd Vendida', 'Valor Total'],
            linhas_prod
        )

        # Top Clientes
        linhas_cli = []
        for idx, item in enumerate(top_clientes, 1):
            nome, vendas_count, valor = item
            linhas_cli.append([idx, nome, int(vendas_count or 0), float(valor or 0)])

        criar_planilha_simples(
            wb,
            'Top Clientes',
            'Top 5 Clientes',
            ['#', 'Cliente', 'Vendas', 'Valor Total'],
            linhas_cli
        )

        # Salvar em buffer e retornar HTTP response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'relatorio_desempenho_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_consolidado_excel(self, filtros):
        """Gera relatório consolidado em Excel com visão geral de todas as operações"""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Exportação para Excel requer a biblioteca openpyxl. Execute: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        from django.db import connection
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # ===== COLETAR DADOS =====
        # VENDAS
        sql_vendas = """
            SELECT 
                COUNT(v.id_venda) as total_vendas,
                COALESCE(SUM((SELECT SUM(vi.valor_total) FROM venda_itens vi WHERE vi.id_venda = v.id_venda)), 0) as valor_total,
                COUNT(DISTINCT v.id_cliente) as clientes_distintos
            FROM vendas v
            WHERE 1=1
        """
        params_vendas = []
        if filtros.get('dataInicio'):
            sql_vendas += " AND DATE(v.data_documento) >= %s"
            params_vendas.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_vendas += " AND DATE(v.data_documento) <= %s"
            params_vendas.append(filtros['dataFim'])
        
        with connection.cursor() as cursor:
            cursor.execute(sql_vendas, params_vendas)
            total_vendas, valor_vendas, clientes_vendas = cursor.fetchone() or (0, 0, 0)
        
        # COMPRAS
        sql_compras = """
            SELECT 
                COUNT(c.id_compra) as total_compras,
                COALESCE(SUM(c.valor_total_nota), 0) as valor_total,
                COUNT(DISTINCT c.id_fornecedor) as fornecedores_distintos
            FROM compras c
            WHERE 1=1
        """
        params_compras = []
        if filtros.get('dataInicio'):
            sql_compras += " AND DATE(c.data_movimento_entrada) >= %s"
            params_compras.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql_compras += " AND DATE(c.data_movimento_entrada) <= %s"
            params_compras.append(filtros['dataFim'])
        
        with connection.cursor() as cursor:
            cursor.execute(sql_compras, params_compras)
            total_compras, valor_compras, fornecedores_compras = cursor.fetchone() or (0, 0, 0)
        
        # ESTOQUE
        sql_estoque = """
            SELECT 
                COUNT(DISTINCT p.id_produto) as total_produtos,
                COALESCE(SUM(e.quantidade), 0) as quantidade_total,
                COALESCE(SUM(e.quantidade * e.valor_venda), 0) as valor_estoque
            FROM produtos p
            LEFT JOIN estoque e ON p.id_produto = e.id_produto
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql_estoque)
            total_produtos, qtd_estoque, valor_estoque = cursor.fetchone() or (0, 0, 0)
        
        # CLIENTES
        sql_clientes = "SELECT COUNT(*) as total_clientes FROM clientes"
        with connection.cursor() as cursor:
            cursor.execute(sql_clientes)
            clientes_ativos = cursor.fetchone()[0] or 0
        
        # FORNECEDORES
        sql_fornecedores = "SELECT COUNT(*) as total_fornecedores FROM fornecedores"
        with connection.cursor() as cursor:
            cursor.execute(sql_fornecedores)
            fornecedores_ativos = cursor.fetchone()[0] or 0
        
        # DEVOLUÇÕES
        try:
            sql_devolucoes = """
                SELECT 
                    COUNT(d.id_devolucao) as total_devolucoes,
                    COALESCE(SUM(d.valor_devolucao), 0) as valor_devolucoes
                FROM devolucoes d
                WHERE 1=1
            """
            params_dev = []
            if filtros.get('dataInicio'):
                sql_devolucoes += " AND DATE(d.data_devolucao) >= %s"
                params_dev.append(filtros['dataInicio'])
            if filtros.get('dataFim'):
                sql_devolucoes += " AND DATE(d.data_devolucao) <= %s"
                params_dev.append(filtros['dataFim'])
            
            with connection.cursor() as cursor:
                cursor.execute(sql_devolucoes, params_dev)
                total_devolucoes, valor_devolucoes = cursor.fetchone() or (0, 0)
        except Exception:
            total_devolucoes, valor_devolucoes = 0, 0
        
        # TROCAS
        try:
            sql_trocas = """
                SELECT 
                    COUNT(t.id_troca) as total_trocas,
                    COALESCE(SUM(CASE WHEN t.status = 'finalizada' THEN t.valor_devolucao ELSE 0 END), 0) as valor_retorno,
                    COALESCE(SUM(CASE WHEN t.status = 'finalizada' THEN t.valor_substituicao ELSE 0 END), 0) as valor_substituicao
                FROM trocas t
                WHERE 1=1
            """
            params_troc = []
            if filtros.get('dataInicio'):
                sql_trocas += " AND DATE(t.data_troca) >= %s"
                params_troc.append(filtros['dataInicio'])
            if filtros.get('dataFim'):
                sql_trocas += " AND DATE(t.data_troca) <= %s"
                params_troc.append(filtros['dataFim'])
            
            with connection.cursor() as cursor:
                cursor.execute(sql_trocas, params_troc)
                total_trocas, valor_retorno, valor_substituicao = cursor.fetchone() or (0, 0, 0)
        except Exception:
            total_trocas, valor_retorno, valor_substituicao = 0, 0, 0
        
        # COMANDAS
        try:
            sql_comandas = """
                SELECT 
                    COUNT(c.id_comanda) as total_comandas,
                    COALESCE(SUM(c.valor_total), 0) as valor_total,
                    COALESCE(SUM(c.desconto), 0) as desconto_total
                FROM comandas c
                WHERE 1=1
            """
            params_cmd = []
            if filtros.get('dataInicio'):
                sql_comandas += " AND DATE(c.data_abertura) >= %s"
                params_cmd.append(filtros['dataInicio'])
            if filtros.get('dataFim'):
                sql_comandas += " AND DATE(c.data_abertura) <= %s"
                params_cmd.append(filtros['dataFim'])
            
            with connection.cursor() as cursor:
                cursor.execute(sql_comandas, params_cmd)
                total_comandas, valor_comandas, desconto_comandas = cursor.fetchone() or (0, 0, 0)
        except Exception:
            total_comandas, valor_comandas, desconto_comandas = 0, 0, 0
        
        # ===== CRIAR WORKBOOK =====
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Consolidado'
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # TÍTULO
        ws['A1'] = 'RELATÓRIO CONSOLIDADO'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1f4e78', end_color='1f4e78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 25
        
        # Período
        if filtros.get('dataInicio') or filtros.get('dataFim'):
            periodo = f"Período: {filtros.get('dataInicio', 'sem data')} a {filtros.get('dataFim', 'sem data')}"
            ws['A2'] = periodo
            ws['A2'].font = Font(italic=True, size=10)
            ws.merge_cells('A2:B2')
        
        # OPERAÇÕES
        row = 4
        ws['A' + str(row)] = 'OPERAÇÕES'
        ws['A' + str(row)].font = header_font
        ws['A' + str(row)].fill = header_fill
        ws['B' + str(row)].font = header_font
        ws['B' + str(row)].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        operacoes_data = [
            ['📊 Total de Vendas', int(total_vendas or 0)],
            ['💰 Valor Vendas', float(valor_vendas or 0)],
            ['📦 Total de Compras', int(total_compras or 0)],
            ['💸 Valor Compras', float(valor_compras or 0)],
            ['↩️ Devoluções Registradas', int(total_devolucoes or 0)],
            ['💝 Valor de Devoluções', float(valor_devolucoes or 0)],
            ['🔄 Trocas Realizadas', int(total_trocas or 0)],
            ['🔙 Valor Retorno (Trocas)', float(valor_retorno or 0)],
            ['🔜 Valor Substituição (Trocas)', float(valor_substituicao or 0)],
            ['📋 Comandas Fechadas', int(total_comandas or 0)],
            ['💳 Faturamento de Comandas', float(valor_comandas or 0)],
        ]
        
        row += 1
        for label, value in operacoes_data:
            ws['A' + str(row)] = label
            ws['B' + str(row)] = value
            ws['A' + str(row)].border = border
            ws['B' + str(row)].border = border
            ws['B' + str(row)].alignment = Alignment(horizontal='right')
            if isinstance(value, float):
                ws['B' + str(row)].number_format = '#,##0.00'
            row += 1
        
        # PRODUTOS E ESTOQUE
        row += 1
        ws['A' + str(row)] = 'ESTOQUE'
        ws['A' + str(row)].font = header_font
        ws['A' + str(row)].fill = header_fill
        ws['B' + str(row)].font = header_font
        ws['B' + str(row)].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        estoque_data = [
            ['📊 Total de Produtos', int(total_produtos or 0)],
            ['📦 Quantidade em Estoque', int(qtd_estoque or 0)],
            ['💵 Valor Total do Estoque', float(valor_estoque or 0)],
        ]
        
        row += 1
        for label, value in estoque_data:
            ws['A' + str(row)] = label
            ws['B' + str(row)] = value
            ws['A' + str(row)].border = border
            ws['B' + str(row)].border = border
            ws['B' + str(row)].alignment = Alignment(horizontal='right')
            if isinstance(value, float):
                ws['B' + str(row)].number_format = '#,##0.00'
            row += 1
        
        # RELACIONAMENTOS
        row += 1
        ws['A' + str(row)] = 'RELACIONAMENTOS'
        ws['A' + str(row)].font = header_font
        ws['A' + str(row)].fill = header_fill
        ws['B' + str(row)].font = header_font
        ws['B' + str(row)].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        relacionamentos_data = [
            ['👥 Clientes Ativos', int(clientes_ativos or 0)],
            ['🏭 Fornecedores Ativos', int(fornecedores_ativos or 0)],
        ]
        
        row += 1
        for label, value in relacionamentos_data:
            ws['A' + str(row)] = label
            ws['B' + str(row)] = value
            ws['A' + str(row)].border = border
            ws['B' + str(row)].border = border
            ws['B' + str(row)].alignment = Alignment(horizontal='right')
            row += 1
        
        # Ajustar largura
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        
        # Salvar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_consolidado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def gerar_relatorio_comissoes_excel(self, filtros):
        """Gera relatório de comissões em Excel"""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Exportação para Excel requer a biblioteca openpyxl. Execute: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        from django.db import connection
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # Remove a planilha padrão
        wb.remove(wb.active)
        
        # ===== ABA 1: RESUMO DE COMISSÕES =====
        ws_resumo = wb.create_sheet('Resumo Comissões')
        
        # Títulos
        ws_resumo['A1'] = 'RELATÓRIO DE COMISSÕES POR VENDEDOR'
        ws_resumo['A1'].font = Font(size=14, bold=True)
        ws_resumo.merge_cells('A1:E1')
        
        # Filtros aplicados
        row = 3
        if filtros.get('dataInicio') or filtros.get('dataFim') or filtros.get('vendedor'):
            ws_resumo[f'A{row}'] = 'Filtros Aplicados:'
            ws_resumo[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if filtros.get('dataInicio'):
                ws_resumo[f'A{row}'] = f"Data Início: {filtros['dataInicio']}"
                row += 1
            if filtros.get('dataFim'):
                ws_resumo[f'A{row}'] = f"Data Fim: {filtros['dataFim']}"
                row += 1
            if filtros.get('vendedor'):
                ws_resumo[f'A{row}'] = f"Vendedor ID: {filtros['vendedor']}"
                row += 1
            row += 1
        
        # Query de comissões
        sql_comissoes = """
            SELECT 
                v.id_vendedor,
                v.nome as vendedor_nome,
                v.percentual_comissao,
                COUNT(DISTINCT vd.id_venda) as total_vendas,
                COALESCE(SUM(vd.valor_total), 0) as valor_total_vendas,
                COALESCE(SUM(vd.valor_total * (v.percentual_comissao / 100)), 0) as comissao_vendas,
                COUNT(DISTINCT CASE WHEN fc.status_conta = 'Paga' THEN fc.id_conta END) as total_contas_recebidas,
                COALESCE(SUM(
                    CASE WHEN fc.status_conta = 'Paga' 
                    THEN fc.valor_liquidado 
                    ELSE 0 END
                ), 0) as valor_recebido,
                COALESCE(SUM(
                    CASE WHEN fc.status_conta = 'Paga' 
                    THEN fc.valor_liquidado * (v.percentual_comissao / 100)
                    ELSE 0 END
                ), 0) as comissao_recebimentos
            FROM vendedores v
            LEFT JOIN vendas vd ON v.id_vendedor = vd.id_vendedor1
            LEFT JOIN financeiro_contas fc ON vd.id_venda = fc.id_venda_origem 
                AND fc.tipo_conta = 'Receber'
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql_comissoes += " AND DATE(vd.data_documento) >= %s"
            params.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_comissoes += " AND DATE(vd.data_documento) <= %s"
            params.append(filtros['dataFim'])
        
        if filtros.get('vendedor'):
            sql_comissoes += " AND v.id_vendedor = %s"
            params.append(filtros['vendedor'])
        
        sql_comissoes += """
            GROUP BY v.id_vendedor, v.nome, v.percentual_comissao
            HAVING total_vendas > 0 OR total_contas_recebidas > 0
            ORDER BY (comissao_vendas + comissao_recebimentos) DESC
        """
        
        with connection.cursor() as cursor:
            cursor.execute(sql_comissoes, params)
            vendedores_comissoes = cursor.fetchall()
        
        # Cabeçalho da tabela resumo
        header_fill = PatternFill(start_color='F57C00', end_color='F57C00', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        ws_resumo[f'A{row}'] = 'Vendedor'
        ws_resumo[f'B{row}'] = '% Comissão'
        ws_resumo[f'C{row}'] = 'Vendas'
        ws_resumo[f'D{row}'] = 'Valor Vendido'
        ws_resumo[f'E{row}'] = 'Comissão Vendas'
        ws_resumo[f'F{row}'] = 'Contas Recebidas'
        ws_resumo[f'G{row}'] = 'Valor Recebido'
        ws_resumo[f'H{row}'] = 'Comissão Recebimentos'
        ws_resumo[f'I{row}'] = 'Total Comissões'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws_resumo[f'{col}{row}'].fill = header_fill
            ws_resumo[f'{col}{row}'].font = header_font
            ws_resumo[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Dados
        for vend in vendedores_comissoes:
            id_vend, nome, perc_comissao, qtd_vendas, valor_vendas, comissao_vendas, qtd_recebidas, valor_recebido, comissao_receb = vend
            
            ws_resumo[f'A{row}'] = nome
            ws_resumo[f'B{row}'] = f"{float(perc_comissao or 0):.1f}%"
            ws_resumo[f'C{row}'] = int(qtd_vendas or 0)
            ws_resumo[f'D{row}'] = float(valor_vendas or 0)
            ws_resumo[f'E{row}'] = float(comissao_vendas or 0)
            ws_resumo[f'F{row}'] = int(qtd_recebidas or 0)
            ws_resumo[f'G{row}'] = float(valor_recebido or 0)
            ws_resumo[f'H{row}'] = float(comissao_receb or 0)
            ws_resumo[f'I{row}'] = float(comissao_vendas or 0) + float(comissao_receb or 0)
            
            # Formatação
            ws_resumo[f'D{row}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'E{row}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'G{row}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'H{row}'].number_format = 'R$ #,##0.00'
            ws_resumo[f'I{row}'].number_format = 'R$ #,##0.00'
            
            row += 1
        
        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 12
        ws_resumo.column_dimensions['C'].width = 10
        ws_resumo.column_dimensions['D'].width = 15
        ws_resumo.column_dimensions['E'].width = 18
        ws_resumo.column_dimensions['F'].width = 15
        ws_resumo.column_dimensions['G'].width = 15
        ws_resumo.column_dimensions['H'].width = 20
        ws_resumo.column_dimensions['I'].width = 18
        
        # ===== ABA 2: DETALHAMENTO DE VENDAS =====
        ws_vendas = wb.create_sheet('Vendas Detalhadas')
        
        ws_vendas['A1'] = 'DETALHAMENTO DAS VENDAS POR VENDEDOR'
        ws_vendas['A1'].font = Font(size=14, bold=True)
        ws_vendas.merge_cells('A1:F1')
        
        # Query vendas detalhadas
        sql_vendas_detalhadas = """
            SELECT 
                v.nome as vendedor_nome,
                v.percentual_comissao,
                vd.id_venda,
                vd.numero_documento,
                DATE_FORMAT(vd.data_documento, '%d/%m/%Y %H:%i') as data_venda,
                c.nome_razao_social as cliente_nome,
                vd.valor_total,
                (vd.valor_total * (v.percentual_comissao / 100)) as comissao_venda
            FROM vendedores v
            INNER JOIN vendas vd ON v.id_vendedor = vd.id_vendedor1
            LEFT JOIN clientes c ON vd.id_cliente = c.id_cliente
            WHERE 1=1
        """
        
        params_detalhadas = []
        
        if filtros.get('dataInicio'):
            sql_vendas_detalhadas += " AND DATE(vd.data_documento) >= %s"
            params_detalhadas.append(filtros['dataInicio'])
        
        if filtros.get('dataFim'):
            sql_vendas_detalhadas += " AND DATE(vd.data_documento) <= %s"
            params_detalhadas.append(filtros['dataFim'])
        
        if filtros.get('vendedor'):
            sql_vendas_detalhadas += " AND v.id_vendedor = %s"
            params_detalhadas.append(filtros['vendedor'])
        
        sql_vendas_detalhadas += " ORDER BY v.nome, vd.data_documento DESC"
        
        with connection.cursor() as cursor:
            cursor.execute(sql_vendas_detalhadas, params_detalhadas)
            vendas_detalhadas = cursor.fetchall()
        
        row = 3
        
        # Cabeçalho
        ws_vendas[f'A{row}'] = 'Vendedor'
        ws_vendas[f'B{row}'] = '% Com.'
        ws_vendas[f'C{row}'] = 'Nº Documento'
        ws_vendas[f'D{row}'] = 'Data'
        ws_vendas[f'E{row}'] = 'Cliente'
        ws_vendas[f'F{row}'] = 'Valor Venda'
        ws_vendas[f'G{row}'] = 'Comissão'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_vendas[f'{col}{row}'].fill = header_fill
            ws_vendas[f'{col}{row}'].font = header_font
            ws_vendas[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Dados
        for venda in vendas_detalhadas:
            vendedor_nome, perc_comissao, id_venda, num_doc, data_venda, cliente_nome, valor_total, comissao = venda
            
            ws_vendas[f'A{row}'] = vendedor_nome
            ws_vendas[f'B{row}'] = f"{float(perc_comissao or 0):.1f}%"
            ws_vendas[f'C{row}'] = num_doc or f'#{id_venda}'
            ws_vendas[f'D{row}'] = data_venda
            ws_vendas[f'E{row}'] = cliente_nome or '-'
            ws_vendas[f'F{row}'] = float(valor_total or 0)
            ws_vendas[f'G{row}'] = float(comissao or 0)
            
            ws_vendas[f'F{row}'].number_format = 'R$ #,##0.00'
            ws_vendas[f'G{row}'].number_format = 'R$ #,##0.00'
            
            row += 1
        
        # Ajustar largura das colunas
        ws_vendas.column_dimensions['A'].width = 30
        ws_vendas.column_dimensions['B'].width = 10
        ws_vendas.column_dimensions['C'].width = 15
        ws_vendas.column_dimensions['D'].width = 18
        ws_vendas.column_dimensions['E'].width = 35
        ws_vendas.column_dimensions['F'].width = 15
        ws_vendas.column_dimensions['G'].width = 15
        
        # Salvar arquivo
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_comissoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    # ============================================================================
    # RELATÓRIOS DE DEVOLUÇÕES
    # ============================================================================

    def gerar_relatorio_devolucoes(self, filtros):
        """Gera relatório de devoluções de vendas"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Devoluções',
            'Análise de devoluções de vendas e impacto financeiro'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('cliente'):
            filtros_text.append(f"Cliente ID: {filtros['cliente']}")
        if filtros.get('status') and filtros['status'] != 'todos':
            filtros_text.append(f"Status: {filtros['status']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar devoluções do banco de dados
        sql = """
            SELECT 
                d.id_devolucao,
                d.numero_devolucao,
                d.data_devolucao,
                d.tipo,
                d.id_cliente,
                d.id_venda,
                d.motivo,
                d.valor_total_devolucao,
                d.status,
                d.gerar_credito,
                c.nome_razao_social as cliente_nome
            FROM devolucoes d
            LEFT JOIN clientes c ON d.id_cliente = c.id_cliente
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(d.data_devolucao) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(d.data_devolucao) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('cliente'):
            sql += " AND d.id_cliente = %s"
            params.append(filtros['cliente'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND d.status = %s"
            params.append(filtros['status'])
        
        sql += " ORDER BY d.data_devolucao DESC LIMIT 100"
        
        # Executar query
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                devolucoes = cursor.fetchall()
        except Exception as e:
            print(f"Erro ao buscar devoluções: {e}")
            devolucoes = []
        
        # Criar tabela com dados
        data = [['#', 'Nº Devolução', 'Data', 'Tipo', 'Cliente', 'Valor', 'Status', 'Crédito']]
        total_devolucoes = 0
        total_valor = 0
        
        for idx, dev in enumerate(devolucoes, 1):
            id_dev, numero_dev, data_dev, tipo, id_cli, id_vend, motivo, valor, status, gerar_cred, cliente = dev
            
            data_fmt = data_dev.strftime('%d/%m/%Y') if data_dev else '-'
            tipo_fmt = 'Venda' if tipo == 'venda' else 'Compra'
            cliente_fmt = cliente or 'N/A'
            valor_fmt = f'R$ {float(valor or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            status_fmt = status or 'Pendente'
            credito_fmt = 'Sim' if gerar_cred else 'Não'
            
            data.append([
                str(idx), 
                numero_dev or f'DEV{id_dev}', 
                data_fmt, 
                tipo_fmt,
                cliente_fmt, 
                valor_fmt, 
                status_fmt,
                credito_fmt
            ])
            total_devolucoes += 1
            total_valor += float(valor or 0)
        
        if len(data) == 1:
            data.append(['-', '-', '-', '-', 'Nenhuma devolução encontrada', '-', '-', '-'])
        
        table = Table(data, colWidths=[0.8*cm, 2*cm, 2.2*cm, 1.5*cm, 5*cm, 2.3*cm, 1.8*cm, 1.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c62828')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ffebee')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
            ('FONTNAME', (5, 1), (5, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Resumo
        total_formatado = f'R$ {total_valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        elements.append(Paragraph(f'<b>Total de Devoluções:</b> {total_devolucoes}', styles['Normal']))
        elements.append(Paragraph(f'<b>Valor Total Devolvido:</b> {total_formatado}', styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_devolucoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def gerar_relatorio_devolucoes_excel(self, filtros):
        """Gera relatório de devoluções em Excel"""
        from django.db import connection
        
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Biblioteca openpyxl não instalada'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Buscar devoluções
        sql = """
            SELECT 
                d.numero_devolucao,
                d.data_devolucao,
                d.tipo,
                c.nome_razao_social,
                d.motivo,
                d.valor_total_devolucao,
                d.status,
                d.gerar_credito
            FROM devolucoes d
            LEFT JOIN clientes c ON d.id_cliente = c.id_cliente
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(d.data_devolucao) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(d.data_devolucao) <= %s"
            params.append(filtros['dataFim'])
        
        sql += " ORDER BY d.data_devolucao DESC LIMIT 1000"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                devolucoes = cursor.fetchall()
        except:
            devolucoes = []
        
        # Preparar dados
        dados = []
        total = 0
        for dev in devolucoes:
            numero, data, tipo, cliente, motivo, valor, status, gerar_cred = dev
            data_fmt = data.strftime('%d/%m/%Y') if data else '-'
            valor_num = float(valor or 0)
            total += valor_num
            dados.append([
                numero or '-',
                data_fmt,
                'Venda' if tipo == 'venda' else 'Compra',
                cliente or 'N/A',
                motivo or '-',
                valor_num,
                status or 'Pendente',
                'Sim' if gerar_cred else 'Não'
            ])
        
        dados.append(['', '', '', '', 'TOTAL', total, '', ''])
        
        wb = self.criar_excel_base(
            'Relatório de Devoluções',
            ['Nº Devolução', 'Data', 'Tipo', 'Cliente', 'Motivo', 'Valor', 'Status', 'Crédito?'],
            dados,
            'Devoluções'
        )
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_devolucoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ============================================================================
    # RELATÓRIOS DE TROCAS
    # ============================================================================

    def gerar_relatorio_trocas(self, filtros):
        """Gera relatório de trocas de produtos"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Trocas',
            'Análise de trocas de produtos e movimentações'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('status') and filtros['status'] != 'todos':
            filtros_text.append(f"Status: {filtros['status']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar trocas do banco de dados
        sql = """
            SELECT 
                t.id_troca,
                t.data_troca,
                t.id_venda_original,
                t.id_cliente,
                t.valor_total_retorno,
                t.valor_total_substituicao,
                t.status,
                t.observacao,
                c.nome_razao_social as cliente_nome
            FROM trocas t
            LEFT JOIN clientes c ON t.id_cliente = c.id_cliente
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(t.data_troca) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(t.data_troca) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND t.status = %s"
            params.append(filtros['status'])
        
        sql += " ORDER BY t.data_troca DESC LIMIT 100"
        
        # Executar query
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                trocas = cursor.fetchall()
        except Exception as e:
            print(f"Erro ao buscar trocas: {e}")
            trocas = []
        
        # Criar tabela com dados
        data = [['#', 'Data', 'Venda Original', 'Cliente', 'Valor Retorno', 'Valor Substituição', 'Diferença', 'Status']]
        total_trocas = 0
        total_retorno = 0
        total_substituicao = 0
        
        for idx, troca in enumerate(trocas, 1):
            id_troca, data_troca, id_venda, id_cli, val_ret, val_sub, status, obs, cliente = troca
            
            data_fmt = data_troca.strftime('%d/%m/%Y') if data_troca else '-'
            cliente_fmt = cliente or 'N/A'
            val_ret_num = float(val_ret or 0)
            val_sub_num = float(val_sub or 0)
            diferenca = val_sub_num - val_ret_num
            
            val_ret_fmt = f'R$ {val_ret_num:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            val_sub_fmt = f'R$ {val_sub_num:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            dif_fmt = f'R$ {diferenca:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            
            data.append([
                str(idx),
                data_fmt,
                str(id_venda),
                cliente_fmt,
                val_ret_fmt,
                val_sub_fmt,
                dif_fmt,
                status or 'Pendente'
            ])
            total_trocas += 1
            total_retorno += val_ret_num
            total_substituicao += val_sub_num
        
        if len(data) == 1:
            data.append(['-', '-', '-', 'Nenhuma troca encontrada', '-', '-', '-', '-'])
        
        table = Table(data, colWidths=[0.7*cm, 2*cm, 2*cm, 4*cm, 2*cm, 2.3*cm, 2*cm, 1.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6a1b9a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3e5f5')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (6, -1), 'RIGHT'),
            ('FONTNAME', (4, 1), (6, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Resumo
        total_ret_fmt = f'R$ {total_retorno:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        total_sub_fmt = f'R$ {total_substituicao:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        dif_total = total_substituicao - total_retorno
        
        elements.append(Paragraph(f'<b>Total de Trocas:</b> {total_trocas}', styles['Normal']))
        elements.append(Paragraph(f'<b>Valor Total Retornado:</b> {total_ret_fmt}', styles['Normal']))
        elements.append(Paragraph(f'<b>Valor Total em Substituições:</b> {total_sub_fmt}', styles['Normal']))
        
        if dif_total >= 0:
            dif_fmt = f'R$ {dif_total:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            elements.append(Paragraph(f'<b>Diferença (a Cobrar):</b> {dif_fmt}', styles['Normal']))
        else:
            dif_fmt = f'R$ {abs(dif_total):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            elements.append(Paragraph(f'<b>Diferença (Crédito):</b> {dif_fmt}', styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_trocas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def gerar_relatorio_trocas_excel(self, filtros):
        """Gera relatório de trocas em Excel"""
        from django.db import connection
        
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Biblioteca openpyxl não instalada'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Buscar trocas
        sql = """
            SELECT 
                t.data_troca,
                t.id_venda_original,
                c.nome_razao_social,
                t.valor_total_retorno,
                t.valor_total_substituicao,
                (t.valor_total_substituicao - t.valor_total_retorno) as diferenca,
                t.status
            FROM trocas t
            LEFT JOIN clientes c ON t.id_cliente = c.id_cliente
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(t.data_troca) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(t.data_troca) <= %s"
            params.append(filtros['dataFim'])
        
        sql += " ORDER BY t.data_troca DESC LIMIT 1000"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                trocas = cursor.fetchall()
        except:
            trocas = []
        
        # Preparar dados
        dados = []
        for troca in trocas:
            data, id_venda, cliente, val_ret, val_sub, dif, status = troca
            data_fmt = data.strftime('%d/%m/%Y') if data else '-'
            dados.append([
                data_fmt,
                str(id_venda),
                cliente or 'N/A',
                float(val_ret or 0),
                float(val_sub or 0),
                float(dif or 0),
                status or 'Pendente'
            ])
        
        wb = self.criar_excel_base(
            'Relatório de Trocas',
            ['Data', 'Venda Original', 'Cliente', 'Valor Retorno', 'Valor Substituição', 'Diferença', 'Status'],
            dados,
            'Trocas'
        )
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_trocas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ============================================================================
    # RELATÓRIOS DE COMANDAS
    # ============================================================================

    def gerar_relatorio_comandas(self, filtros):
        """Gera relatório de comandas (pedidos)"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de Comandas',
            'Análise de pedidos e vendas em comandas'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('status') and filtros['status'] != 'todos':
            filtros_text.append(f"Status: {filtros['status']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar comandas do banco de dados
        sql = """
            SELECT 
                c.numero,
                c.data_abertura,
                c.data_fechamento,
                m.numero as mesa_numero,
                cl.nome_razao_social as cliente_nome,
                u.first_name as garcom_nome,
                c.forma_pagamento,
                c.subtotal,
                c.desconto,
                c.taxa_servico,
                c.total,
                c.status
            FROM comandas c
            LEFT JOIN mesas m ON c.mesa_id = m.id
            LEFT JOIN clientes cl ON c.cliente_id = cl.id_cliente
            LEFT JOIN auth_user u ON c.garcom_id = u.id
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(c.data_abertura) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(c.data_abertura) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND c.status = %s"
            params.append(filtros['status'])
        
        sql += " ORDER BY c.data_abertura DESC LIMIT 100"
        
        # Executar query
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                comandas = cursor.fetchall()
        except Exception as e:
            print(f"Erro ao buscar comandas: {e}")
            comandas = []
        
        # Criar tabela com dados
        data = [['#', 'Nº Comanda', 'Data', 'Mesa', 'Garçom', 'Cliente', 'Subtotal', 'Desconto', 'Taxa', 'Total', 'Status']]
        total_comandas = 0
        total_valor = 0
        total_desconto = 0
        
        for idx, cmd in enumerate(comandas, 1):
            numero, data_ab, data_fech, mesa_num, cliente, garcom, forma_pag, subtotal, desc, taxa, total, status = cmd
            
            data_fmt = data_ab.strftime('%d/%m/%Y %H:%M') if data_ab else '-'
            mesa_fmt = f'Mesa {mesa_num}' if mesa_num else 'S/ Mesa'
            cliente_fmt = cliente or '-'
            garcom_fmt = garcom or '-'
            
            subtotal_fmt = f'R$ {float(subtotal or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            desc_fmt = f'R$ {float(desc or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            taxa_fmt = f'R$ {float(taxa or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            total_fmt = f'R$ {float(total or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            
            data.append([
                str(idx),
                numero or '-',
                data_fmt,
                mesa_fmt,
                garcom_fmt,
                cliente_fmt,
                subtotal_fmt,
                desc_fmt,
                taxa_fmt,
                total_fmt,
                status or 'Aberta'
            ])
            
            total_comandas += 1
            total_valor += float(total or 0)
            total_desconto += float(desc or 0)
        
        if len(data) == 1:
            data.append(['-', '-', '-', '-', '-', 'Nenhuma comanda encontrada', '-', '-', '-', '-', '-'])
        
        table = Table(data, colWidths=[0.6*cm, 1.8*cm, 2.2*cm, 1.8*cm, 1.8*cm, 3*cm, 1.5*cm, 1.5*cm, 1.3*cm, 1.5*cm, 1.3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e3f2fd')]),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (6, 1), (9, -1), 'RIGHT'),
            ('FONTNAME', (9, 1), (9, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 1*cm))
        
        # Resumo
        total_fmt = f'R$ {total_valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        desc_total_fmt = f'R$ {total_desconto:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        
        elements.append(Paragraph(f'<b>Total de Comandas Fechadas:</b> {total_comandas}', styles['Normal']))
        elements.append(Paragraph(f'<b>Faturamento Total:</b> {total_fmt}', styles['Normal']))
        elements.append(Paragraph(f'<b>Desconto Total Concedido:</b> {desc_total_fmt}', styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'relatorio_comandas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def gerar_relatorio_comandas_excel(self, filtros):
        """Gera relatório de comandas em Excel"""
        from django.db import connection
        
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Biblioteca openpyxl não instalada'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Buscar comandas
        sql = """
            SELECT 
                c.numero,
                c.data_abertura,
                m.numero as mesa_numero,
                cl.nome_razao_social,
                u.first_name as garcom_nome,
                c.forma_pagamento,
                c.subtotal,
                c.desconto,
                c.taxa_servico,
                c.total,
                c.status
            FROM comandas c
            LEFT JOIN mesas m ON c.mesa_id = m.id
            LEFT JOIN clientes cl ON c.cliente_id = cl.id_cliente
            LEFT JOIN auth_user u ON c.garcom_id = u.id
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(c.data_abertura) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(c.data_abertura) <= %s"
            params.append(filtros['dataFim'])
        
        sql += " ORDER BY c.data_abertura DESC LIMIT 1000"
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                comandas = cursor.fetchall()
        except:
            comandas = []
        
        # Preparar dados
        dados = []
        total_faturamento = 0
        for cmd in comandas:
            numero, data, mesa, cliente, garcom, forma_pag, subtotal, desc, taxa, total, status = cmd
            data_fmt = data.strftime('%d/%m/%Y %H:%M') if data else '-'
            mesa_fmt = f'Mesa {mesa}' if mesa else 'S/ Mesa'
            total_num = float(total or 0)
            total_faturamento += total_num
            dados.append([
                numero or '-',
                data_fmt,
                mesa_fmt,
                garcom or '-',
                cliente or '-',
                forma_pag or '-',
                float(subtotal or 0),
                float(desc or 0),
                float(taxa or 0),
                total_num,
                status or 'Aberta'
            ])
        
        dados.append(['', '', '', '', '', 'TOTAL', '', '', '', total_faturamento, ''])
        
        wb = self.criar_excel_base(
            'Relatório de Comandas',
            ['Nº Comanda', 'Data', 'Mesa', 'Garçom', 'Cliente', 'Forma Pagamento', 'Subtotal', 'Desconto', 'Taxa', 'Total', 'Status'],
            dados,
            'Comandas'
        )
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_comandas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def gerar_excel_simples(self, tipo, filtros):
        """Método genérico para gerar Excel de outros relatórios"""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Exportação para Excel requer a biblioteca openpyxl. Execute: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Por enquanto, retorna mensagem
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo.capitalize()
        ws['A1'] = f'Relatório de {tipo.capitalize()}'
        ws['A2'] = 'Exportação em desenvolvimento...'
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f'relatorio_{tipo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ========== FICHA DE PRODUTO (MOVIMENTAÇÕES) ==========
    
    @action(detail=False, methods=['get'], url_path='ficha-produto')
    def ficha_produto(self, request):
        """
        Retorna todas as movimentações de um produto específico
        Includes: vendas, compras, comandas, ordem de serviço
        """
        from django.db import connection
        from .models import Produto
        
        produto_id = request.query_params.get('produto_id')
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')
        
        if not produto_id:
            return Response(
                {'error': 'O parâmetro produto_id é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            produto = Produto.objects.get(id_produto=produto_id)
        except Produto.DoesNotExist:
            return Response(
                {'error': 'Produto não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        movimentacoes = []
        
        # ===== VENDAS =====
        sql_vendas = """
            SELECT 
                'VENDA' as tipo,
                v.id_venda as documento_id,
                v.numero_documento as numero,
                v.data_documento as data,
                vi.quantidade,
                vi.valor_unitario,
                vi.valor_total,
                c.nome_razao_social as pessoa,
                o.nome_operacao as operacao,
                'Saída' as tipo_movimento
            FROM venda_itens vi
            INNER JOIN vendas v ON vi.id_venda = v.id_venda
            LEFT JOIN clientes c ON v.id_cliente = c.id_cliente
            LEFT JOIN operacoes o ON v.id_operacao = o.id_operacao
            WHERE vi.id_produto = %s
        """
        params_vendas = [produto_id]
        
        if data_inicio:
            sql_vendas += " AND DATE(v.data_documento) >= %s"
            params_vendas.append(data_inicio)
        if data_fim:
            sql_vendas += " AND DATE(v.data_documento) <= %s"
            params_vendas.append(data_fim)
        
        sql_vendas += " ORDER BY v.data_documento DESC"
        
        # ===== COMPRAS =====
        sql_compras = """
            SELECT 
                'COMPRA' as tipo,
                c.id_compra as documento_id,
                c.numero_nota as numero,
                COALESCE(c.data_entrada, c.data_movimento_entrada) as data,
                ci.quantidade,
                ci.valor_unitario,
                ci.valor_total,
                f.nome_razao_social as pessoa,
                o.nome_operacao as operacao,
                'Entrada' as tipo_movimento
            FROM compra_itens ci
            INNER JOIN compras c ON ci.id_compra = c.id_compra
            LEFT JOIN fornecedores f ON c.id_fornecedor = f.id_fornecedor
            LEFT JOIN operacoes o ON c.id_operacao = o.id_operacao
            WHERE ci.id_produto = %s
        """
        params_compras = [produto_id]
        
        if data_inicio:
            sql_compras += " AND DATE(COALESCE(c.data_entrada, c.data_movimento_entrada)) >= %s"
            params_compras.append(data_inicio)
        if data_fim:
            sql_compras += " AND DATE(COALESCE(c.data_entrada, c.data_movimento_entrada)) <= %s"
            params_compras.append(data_fim)
        
        sql_compras += " ORDER BY COALESCE(c.data_entrada, c.data_movimento_entrada) DESC"
        
        # ===== COMANDAS =====
        sql_comandas = """
            SELECT 
                'COMANDA' as tipo,
                c.id_comanda as documento_id,
                c.numero_comanda as numero,
                c.data_abertura as data,
                ci.quantidade,
                ci.preco_unitario as valor_unitario,
                (ci.quantidade * ci.preco_unitario) as valor_total,
                cl.nome_razao_social as pessoa,
                'Comanda' as operacao,
                'Saída' as tipo_movimento
            FROM comanda_itens ci
            INNER JOIN comandas c ON ci.id_comanda = c.id_comanda
            LEFT JOIN clientes cl ON c.id_cliente = cl.id_cliente
            WHERE ci.id_produto = %s
        """
        params_comandas = [produto_id]
        
        if data_inicio:
            sql_comandas += " AND DATE(c.data_abertura) >= %s"
            params_comandas.append(data_inicio)
        if data_fim:
            sql_comandas += " AND DATE(c.data_abertura) <= %s"
            params_comandas.append(data_fim)
        
        sql_comandas += " ORDER BY c.data_abertura DESC"
        
        # ===== ORDEM DE SERVIÇO =====
        sql_os = """
            SELECT 
                'OS' as tipo,
                os.id_os as documento_id,
                os.numero_os as numero,
                os.data_abertura as data,
                osi.quantidade,
                osi.valor_unitario,
                osi.valor_total,
                c.nome_razao_social as pessoa,
                'Ordem de Serviço' as operacao,
                'Saída' as tipo_movimento
            FROM ordem_servico_itens osi
            INNER JOIN ordem_servico os ON osi.id_os = os.id_os
            LEFT JOIN clientes c ON os.id_cliente = c.id_cliente
            WHERE osi.id_produto = %s
        """
        params_os = [produto_id]
        
        if data_inicio:
            sql_os += " AND DATE(os.data_abertura) >= %s"
            params_os.append(data_inicio)
        if data_fim:
            sql_os += " AND DATE(os.data_abertura) <= %s"
            params_os.append(data_fim)
        
        sql_os += " ORDER BY os.data_abertura DESC"
        
        # Executar queries
        with connection.cursor() as cursor:
            # Vendas
            try:
                cursor.execute(sql_vendas, params_vendas)
                vendas = cursor.fetchall()
                for v in vendas:
                    movimentacoes.append({
                        'tipo': v[0],
                        'documento_id': v[1],
                        'numero': v[2],
                        'data': v[3].strftime('%d/%m/%Y %H:%M') if v[3] else '',
                        'quantidade': float(v[4]) if v[4] else 0,
                        'valor_unitario': float(v[5]) if v[5] else 0,
                        'valor_total': float(v[6]) if v[6] else 0,
                        'pessoa': v[7] or 'Não informado',
                        'operacao': v[8] or 'Sem operação',
                        'tipo_movimento': v[9]
                    })
            except Exception as e:
                logger.error(f'Erro ao buscar vendas: {e}')
            
            # Compras
            try:
                cursor.execute(sql_compras, params_compras)
                compras = cursor.fetchall()
                for c in compras:
                    movimentacoes.append({
                        'tipo': c[0],
                        'documento_id': c[1],
                        'numero': c[2],
                        'data': c[3].strftime('%d/%m/%Y') if c[3] else '',
                        'quantidade': float(c[4]) if c[4] else 0,
                        'valor_unitario': float(c[5]) if c[5] else 0,
                        'valor_total': float(c[6]) if c[6] else 0,
                        'pessoa': c[7] or 'Não informado',
                        'operacao': c[8] or 'Sem operação',
                        'tipo_movimento': c[9]
                    })
            except Exception as e:
                logger.error(f'Erro ao buscar compras: {e}')
            
            # Comandas
            try:
                cursor.execute(sql_comandas, params_comandas)
                comandas = cursor.fetchall()
                for cmd in comandas:
                    movimentacoes.append({
                        'tipo': cmd[0],
                        'documento_id': cmd[1],
                        'numero': cmd[2],
                        'data': cmd[3].strftime('%d/%m/%Y %H:%M') if cmd[3] else '',
                        'quantidade': float(cmd[4]) if cmd[4] else 0,
                        'valor_unitario': float(cmd[5]) if cmd[5] else 0,
                        'valor_total': float(cmd[6]) if cmd[6] else 0,
                        'pessoa': cmd[7] or 'Não informado',
                        'operacao': cmd[8],
                        'tipo_movimento': cmd[9]
                    })
            except Exception as e:
                logger.error(f'Erro ao buscar comandas: {e}')
            
            # Ordem de Serviço
            try:
                cursor.execute(sql_os, params_os)
                ordens = cursor.fetchall()
                for os in ordens:
                    movimentacoes.append({
                        'tipo': os[0],
                        'documento_id': os[1],
                        'numero': os[2],
                        'data': os[3].strftime('%d/%m/%Y %H:%M') if os[3] else '',
                        'quantidade': float(os[4]) if os[4] else 0,
                        'valor_unitario': float(os[5]) if os[5] else 0,
                        'valor_total': float(os[6]) if os[6] else 0,
                        'pessoa': os[7] or 'Não informado',
                        'operacao': os[8],
                        'tipo_movimento': os[9]
                    })
            except Exception as e:
                logger.error(f'Erro ao buscar ordens de serviço: {e}')
        
        # Ordenar todas as movimentações por data (mais recente primeiro)
        movimentacoes.sort(key=lambda x: x['data'], reverse=True)
        
        # Calcular totais
        total_entradas = sum(m['quantidade'] for m in movimentacoes if m['tipo_movimento'] == 'Entrada')
        total_saidas = sum(m['quantidade'] for m in movimentacoes if m['tipo_movimento'] == 'Saída')
        saldo = total_entradas - total_saidas
        
        return Response({
            'produto': {
                'id': produto.id_produto,
                'codigo': produto.codigo_produto,
                'nome': produto.nome_produto,
                'estoque_atual': float(produto.estoque_atual) if hasattr(produto, 'estoque_atual') else 0,
            },
            'movimentacoes': movimentacoes,
            'resumo': {
                'total_movimentacoes': len(movimentacoes),
                'total_entradas': total_entradas,
                'total_saidas': total_saidas,
                'saldo': saldo
            }
        })

    @action(detail=False, methods=['get'], url_path='lucratividade')
    def relatorio_lucratividade(self, request):
        """
        Relatório de Lucratividade por Produto
        Calcula: Receita, Custo, Lucro Bruto e Margem de Lucro
        
        Parâmetros:
        - data_inicial: Data inicial (formato: YYYY-MM-DD)
        - data_final: Data final (formato: YYYY-MM-DD)
        - id_produto: ID do produto (opcional)
        - id_grupo: ID do grupo de produtos (opcional)
        - formato: 'json' ou 'excel' (padrão: json)
        """
        from api.models import VendaItem, Estoque, Produto, GrupoProduto
        from django.db.models import Sum, F, Q, DecimalField, Case, When
        from django.db.models.functions import Coalesce
        from decimal import Decimal
        
        # Parâmetros
        data_inicial = request.query_params.get('data_inicial')
        data_final = request.query_params.get('data_final')
        id_produto = request.query_params.get('id_produto')
        id_grupo = request.query_params.get('id_grupo')
        formato = request.query_params.get('formato', 'json')
        
        # Filtros base
        filtros = Q()
        if data_inicial:
            filtros &= Q(id_venda__data_documento__gte=data_inicial)
        if data_final:
            filtros &= Q(id_venda__data_documento__lte=data_final)
        if id_produto:
            filtros &= Q(id_produto__id_produto=id_produto)
        if id_grupo:
            filtros &= Q(id_produto__id_grupo__id_grupo=id_grupo)
        
        # Buscar itens de venda
        itens = VendaItem.objects.filter(filtros).select_related(
            'id_produto', 'id_produto__id_grupo'
        ).values(
            'id_produto__id_produto',
            'id_produto__codigo_produto',
            'id_produto__nome_produto',
            'id_produto__id_grupo__nome_grupo'
        ).annotate(
            quantidade_vendida=Sum('quantidade'),
            receita_total=Sum(F('valor_unitario') * F('quantidade'), output_field=DecimalField())
        )
        
        # Calcular lucro para cada produto
        resultado = []
        for item in itens:
            try:
                # Buscar custo médio do estoque
                estoque = Estoque.objects.filter(
                    id_produto__id_produto=item['id_produto__id_produto']
                ).first()
                
                custo_medio = estoque.custo_medio if estoque else Decimal('0')
                quantidade = item['quantidade_vendida'] or Decimal('0')
                receita = item['receita_total'] or Decimal('0')
                custo_total = custo_medio * quantidade
                lucro_bruto = receita - custo_total
                margem_lucro = (lucro_bruto / receita * 100) if receita > 0 else Decimal('0')
                
                resultado.append({
                    'id_produto': item['id_produto__id_produto'],
                    'codigo': item['id_produto__codigo_produto'],
                    'nome': item['id_produto__nome_produto'],
                    'grupo': item['id_produto__id_grupo__nome_grupo'] or 'Sem Grupo',
                    'quantidade_vendida': float(quantidade),
                    'receita_total': float(receita),
                    'custo_medio_unitario': float(custo_medio),
                    'custo_total': float(custo_total),
                    'lucro_bruto': float(lucro_bruto),
                    'margem_lucro_percentual': float(margem_lucro)
                })
            except Exception as e:
                logger.error(f'Erro ao calcular lucratividade do produto {item["id_produto__id_produto"]}: {e}')
                continue
        
        # Ordenar por lucro bruto (maior para menor)
        resultado.sort(key=lambda x: x['lucro_bruto'], reverse=True)
        
        # Calcular totais gerais
        totais = {
            'quantidade_total': sum(r['quantidade_vendida'] for r in resultado),
            'receita_total': sum(r['receita_total'] for r in resultado),
            'custo_total': sum(r['custo_total'] for r in resultado),
            'lucro_total': sum(r['lucro_bruto'] for r in resultado),
            'margem_media': (sum(r['lucro_bruto'] for r in resultado) / sum(r['receita_total'] for r in resultado) * 100) if sum(r['receita_total'] for r in resultado) > 0 else 0
        }
        
        if formato == 'excel':
            return self._gerar_excel_lucratividade(resultado, totais, data_inicial, data_final)
        
        return Response({
            'periodo': {
                'data_inicial': data_inicial,
                'data_final': data_final
            },
            'produtos': resultado,
            'totais': totais
        })
    
    @action(detail=False, methods=['get'], url_path='projecao-compra')
    def relatorio_projecao_compra(self, request):
        """
        Relatório de Projeção de Compra
        Lista produtos abaixo do estoque mínimo e sugere quantidade de compra
        
        Parâmetros:
        - id_deposito: ID do depósito (opcional, padrão: todos)
        - id_grupo: ID do grupo de produtos (opcional)
        - somente_abaixo_minimo: true/false (padrão: true)
        - dias_projecao: Número de dias para projeção baseado em vendas (padrão: 30)
        - formato: 'json' ou 'excel' (padrão: json)
        """
        from api.models import Estoque, Produto, VendaItem, Deposito
        from django.db.models import Sum, F, Q, Avg
        from django.utils import timezone
        from datetime import timedelta
        from decimal import Decimal
        
        # Parâmetros
        id_deposito = request.query_params.get('id_deposito')
        id_grupo = request.query_params.get('id_grupo')
        somente_abaixo_minimo = request.query_params.get('somente_abaixo_minimo', 'true').lower() == 'true'
        dias_projecao = int(request.query_params.get('dias_projecao', 30))
        formato = request.query_params.get('formato', 'json')
        
        # Filtros
        filtros = Q(ativo=True)
        if id_deposito:
            filtros &= Q(id_deposito__id_deposito=id_deposito)
        if id_grupo:
            filtros &= Q(id_produto__id_grupo__id_grupo=id_grupo)
        if somente_abaixo_minimo:
            filtros &= Q(quantidade__lte=F('quantidade_minima'))
        
        # Buscar estoques
        estoques = Estoque.objects.filter(filtros).select_related(
            'id_produto', 'id_deposito', 'id_produto__id_grupo'
        )
        
        # Data para cálculo de consumo médio
        data_limite = timezone.now() - timedelta(days=dias_projecao)
        
        resultado = []
        for estoque in estoques:
            try:
                # Calcular consumo médio nos últimos X dias
                vendas = VendaItem.objects.filter(
                    id_produto=estoque.id_produto,
                    id_venda__data_documento__gte=data_limite
                ).aggregate(
                    total_vendido=Sum('quantidade')
                )
                
                quantidade_vendida = vendas['total_vendido'] or Decimal('0')
                consumo_medio_diario = quantidade_vendida / Decimal(str(dias_projecao))
                dias_ate_acabar = (estoque.quantidade / consumo_medio_diario) if consumo_medio_diario > 0 else Decimal('999')
                
                # Sugestão de compra
                # Comprar o suficiente para atingir a quantidade máxima ou 3x o mínimo se máximo não definido
                quantidade_ideal = estoque.quantidade_maxima or (estoque.quantidade_minima * 3)
                quantidade_sugerida = max(quantidade_ideal - estoque.quantidade, Decimal('0'))
                
                # Valor estimado da compra
                valor_estimado = quantidade_sugerida * estoque.valor_ultima_compra if estoque.valor_ultima_compra else Decimal('0')
                
                resultado.append({
                    'id_produto': estoque.id_produto.id_produto,
                    'codigo': estoque.id_produto.codigo_produto,
                    'nome': estoque.id_produto.nome_produto,
                    'grupo': estoque.id_produto.id_grupo.nome_grupo if estoque.id_produto.id_grupo else 'Sem Grupo',
                    'deposito': estoque.id_deposito.nome_deposito,
                    'quantidade_atual': float(estoque.quantidade),
                    'quantidade_minima': float(estoque.quantidade_minima),
                    'quantidade_maxima': float(estoque.quantidade_maxima) if estoque.quantidade_maxima else None,
                    'consumo_medio_diario': float(consumo_medio_diario),
                    'dias_ate_acabar': float(dias_ate_acabar) if dias_ate_acabar < 999 else None,
                    'quantidade_sugerida': float(quantidade_sugerida),
                    'valor_ultima_compra': float(estoque.valor_ultima_compra),
                    'valor_estimado_compra': float(valor_estimado),
                    'status': 'CRÍTICO' if estoque.quantidade <= 0 else 'URGENTE' if estoque.quantidade < estoque.quantidade_minima / 2 else 'BAIXO'
                })
            except Exception as e:
                logger.error(f'Erro ao calcular projeção para produto {estoque.id_produto.id_produto}: {e}')
                continue
        
        # Ordenar por dias até acabar (mais crítico primeiro)
        resultado.sort(key=lambda x: x['dias_ate_acabar'] if x['dias_ate_acabar'] is not None else 999)
        
        # Calcular totais
        totais = {
            'total_produtos': len(resultado),
            'valor_total_estimado': sum(r['valor_estimado_compra'] for r in resultado),
            'produtos_criticos': len([r for r in resultado if r['status'] == 'CRÍTICO']),
            'produtos_urgentes': len([r for r in resultado if r['status'] == 'URGENTE'])
        }
        
        if formato == 'excel':
            return self._gerar_excel_projecao_compra(resultado, totais, dias_projecao)
        
        return Response({
            'configuracao': {
                'dias_projecao': dias_projecao,
                'somente_abaixo_minimo': somente_abaixo_minimo
            },
            'produtos': resultado,
            'totais': totais
        })
    
    def _gerar_excel_lucratividade(self, dados, totais, data_inicial, data_final):
        """Gera arquivo Excel para o relatório de lucratividade"""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Biblioteca openpyxl não disponível'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lucratividade'
        
        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = 'RELATÓRIO DE LUCRATIVIDADE'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Período
        ws.merge_cells('A2:J2')
        periodo = f'Período: {data_inicial or "Início"} até {data_final or "Hoje"}'
        ws['A2'] = periodo
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = ['Código', 'Produto', 'Grupo', 'Qtd Vendida', 'Receita Total', 
                   'Custo Médio', 'Custo Total', 'Lucro Bruto', 'Margem %', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        row = 5
        for item in dados:
            ws.cell(row=row, column=1, value=item['codigo'])
            ws.cell(row=row, column=2, value=item['nome'])
            ws.cell(row=row, column=3, value=item['grupo'])
            ws.cell(row=row, column=4, value=item['quantidade_vendida'])
            ws.cell(row=row, column=5, value=item['receita_total']).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=6, value=item['custo_medio_unitario']).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=7, value=item['custo_total']).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=8, value=item['lucro_bruto']).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=9, value=item['margem_lucro_percentual'] / 100).number_format = '0.00%'
            
            # Status baseado na margem
            margem = item['margem_lucro_percentual']
            status_text = 'Excelente' if margem >= 30 else 'Bom' if margem >= 20 else 'Regular' if margem >= 10 else 'Baixo'
            ws.cell(row=row, column=10, value=status_text)
            row += 1
        
        # Totais
        row += 1
        ws.cell(row=row, column=3, value='TOTAIS:').font = Font(bold=True)
        ws.cell(row=row, column=4, value=totais['quantidade_total']).font = Font(bold=True)
        ws.cell(row=row, column=5, value=totais['receita_total']).font = Font(bold=True)
        ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=7, value=totais['custo_total']).font = Font(bold=True)
        ws.cell(row=row, column=7).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=8, value=totais['lucro_total']).font = Font(bold=True)
        ws.cell(row=row, column=8).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=9, value=totais['margem_media'] / 100).font = Font(bold=True)
        ws.cell(row=row, column=9).number_format = '0.00%'
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=lucratividade_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response
    
    def _gerar_excel_projecao_compra(self, dados, totais, dias_projecao):
        """Gera arquivo Excel para o relatório de projeção de compra"""
        if not EXCEL_AVAILABLE:
            return Response(
                {'error': 'Biblioteca openpyxl não disponível'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Projeção de Compra'
        
        # Título
        ws.merge_cells('A1:L1')
        ws['A1'] = 'RELATÓRIO DE PROJEÇÃO DE COMPRA'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Info
        ws.merge_cells('A2:L2')
        ws['A2'] = f'Projeção baseada em {dias_projecao} dias de venda'
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = ['Status', 'Código', 'Produto', 'Grupo', 'Depósito', 'Qtd Atual', 
                   'Qtd Mínima', 'Consumo/Dia', 'Dias p/ Acabar', 'Qtd Sugerida', 
                   'Vlr Últ. Compra', 'Vlr Estimado']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        row = 5
        for item in dados:
            # Cor de fundo baseada no status
            fill_color = 'FFE6E6' if item['status'] == 'CRÍTICO' else 'FFF4E6' if item['status'] == 'URGENTE' else 'FFFFFF'
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            
            ws.cell(row=row, column=1, value=item['status']).fill = fill
            ws.cell(row=row, column=2, value=item['codigo']).fill = fill
            ws.cell(row=row, column=3, value=item['nome']).fill = fill
            ws.cell(row=row, column=4, value=item['grupo']).fill = fill
            ws.cell(row=row, column=5, value=item['deposito']).fill = fill
            ws.cell(row=row, column=6, value=item['quantidade_atual']).fill = fill
            ws.cell(row=row, column=7, value=item['quantidade_minima']).fill = fill
            ws.cell(row=row, column=8, value=item['consumo_medio_diario']).fill = fill
            ws.cell(row=row, column=9, value=item['dias_ate_acabar'] if item['dias_ate_acabar'] else 'Sem vendas').fill = fill
            ws.cell(row=row, column=10, value=item['quantidade_sugerida']).fill = fill
            ws.cell(row=row, column=11, value=item['valor_ultima_compra']).fill = fill
            ws.cell(row=row, column=11).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=12, value=item['valor_estimado_compra']).fill = fill
            ws.cell(row=row, column=12).number_format = 'R$ #,##0.00'
            row += 1
        
        # Resumo
        row += 2
        ws.cell(row=row, column=1, value='RESUMO:').font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value=f'Total de produtos: {totais["total_produtos"]}')
        row += 1
        ws.cell(row=row, column=1, value=f'Produtos críticos (sem estoque): {totais["produtos_criticos"]}')
        row += 1
        ws.cell(row=row, column=1, value=f'Produtos urgentes (abaixo 50% mín): {totais["produtos_urgentes"]}')
        row += 1
        ws.cell(row=row, column=1, value='Valor total estimado:')
        ws.cell(row=row, column=2, value=totais['valor_total_estimado']).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=projecao_compra_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response
    
    def gerar_relatorio_cte(self, filtros):
        """Gera relatório de CT-e em PDF"""
        from .models import Cliente
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de CT-e',
            'Conhecimentos de Transporte Eletrônicos'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('destinatario'):
            try:
                cliente = Cliente.objects.get(id_cliente=filtros['destinatario'])
                filtros_text.append(f"Destinatário: {cliente.nome_razao_social}")
            except Cliente.DoesNotExist:
                filtros_text.append(f"Destinatário ID: {filtros['destinatario']}")
        if filtros.get('remetente'):
            try:
                cliente = Cliente.objects.get(id_cliente=filtros['remetente'])
                filtros_text.append(f"Remetente: {cliente.nome_razao_social}")
            except Cliente.DoesNotExist:
                filtros_text.append(f"Remetente ID: {filtros['remetente']}")
        if filtros.get('status') and filtros['status'] != 'todos':
            filtros_text.append(f"Status: {filtros['status'].upper()}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar CT-es do banco
        sql = """
            SELECT 
                numero,
                data_emissao,
                remetente_nome,
                destinatario_nome,
                valor_total_servico,
                tipo_servico,
                status
            FROM conhecimento_transporte
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(data_emissao) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(data_emissao) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('destinatario'):
            sql += " AND destinatario_id = %s"
            params.append(filtros['destinatario'])
        if filtros.get('remetente'):
            sql += " AND remetente_id = %s"
            params.append(filtros['remetente'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND status = %s"
            params.append(filtros['status'])
        if filtros.get('numero_c te'):
            sql += " AND numero LIKE %s"
            params.append(f"%{filtros['numero_cte']}%")
        
        sql += " ORDER BY data_emissao DESC LIMIT 100"
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            ctes = cursor.fetchall()
        
        # Criar tabela
        data = [['Número', 'Data', 'Remetente', 'Destinatário', 'Valor', 'Tipo Serviço', 'Status']]
        total_valor = 0
        
        for cte in ctes:
            numero, data_emissao, remetente, destinatario, valor, tipo_servico, status = cte
            data_formatada = data_emissao.strftime('%d/%m/%Y') if data_emissao else '-'
            valor_num = float(valor) if valor else 0
            valor_formatado = f'R$ {valor_num:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            
            data.append([
                str(numero or '-'),
                data_formatada,
                (remetente or 'N/A')[:20],
                (destinatario or 'N/A')[:20],
                valor_formatado,
                (tipo_servico or 'Normal')[:15],
                (status or 'Pendente')[:12]
            ])
            total_valor += valor_num
        
        if len(data) == 1:
            data.append(['-', '-', '-', 'Nenhum CT-e encontrado', '-', '-', '-'])
        
        table = Table(data, colWidths=[2*cm, 2*cm, 3.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1565c0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Totais
        total_formatado = f'R$ {total_valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        elements.append(Paragraph(f"<b>Total de CT-es:</b> {len(ctes)}", styles['Normal']))
        elements.append(Paragraph(f"<b>Valor Total:</b> {total_formatado}", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        return HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    def gerar_relatorio_cte_excel(self, filtros):
        """Gera relatório de CT-e em Excel"""
        if not EXCEL_AVAILABLE:
            return Response({'error': 'Biblioteca openpyxl não disponível'}, status=500)
        
        from django.db import connection
        
        # Buscar dados
        sql = """
            SELECT 
                numero, data_emissao, remetente_nome, destinatario_nome,
                valor_total_servico, tipo_servico, status, chave_acesso
            FROM conhecimento_transporte
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(data_emissao) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(data_emissao) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND status = %s"
            params.append(filtros['status'])
        
        sql += " ORDER BY data_emissao DESC"
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            ctes = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'CT-e'
        
        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = 'RELATÓRIO DE CT-e'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = ['Número', 'Data Emissão', 'Remetente', 'Destinatário', 'Valor Total', 'Tipo Serviço', 'Status', 'Chave Acesso']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1565c0', end_color='1565c0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        row = 4
        total_valor = 0
        for cte in ctes:
            numero, data_emissao, remetente, destinatario, valor, tipo_servico, status, chave = cte
            ws.cell(row=row, column=1, value=numero)
            ws.cell(row=row, column=2, value=data_emissao.strftime('%d/%m/%Y') if data_emissao else '')
            ws.cell(row=row, column=3, value=remetente)
            ws.cell(row=row, column=4, value=destinatario)
            ws.cell(row=row, column=5, value=float(valor) if valor else 0)
            ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=6, value=tipo_servico)
            ws.cell(row=row, column=7, value=status)
            ws.cell(row=row, column=8, value=chave)
            total_valor += float(valor) if valor else 0
            row += 1
        
        # Totais
        row += 1
        ws.cell(row=row, column=4, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_valor).font = Font(bold=True)
        ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 50
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=cte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response
    
    def gerar_relatorio_mdfe(self, filtros):
        """Gera relatório de MDF-e em PDF"""
        from django.db import connection
        
        buffer, doc, elements, styles = self.criar_pdf_base(
            'Relatório de MDF-e',
            'Manifestos Eletrônicos de Documentos Fiscais'
        )
        
        # Mostrar filtros aplicados
        filtros_text = []
        if filtros.get('dataInicio'):
            filtros_text.append(f"Data Início: {filtros['dataInicio']}")
        if filtros.get('dataFim'):
            filtros_text.append(f"Data Fim: {filtros['dataFim']}")
        if filtros.get('status') and filtros['status'] != 'todos':
            filtros_text.append(f"Status: {filtros['status'].upper()}")
        if filtros.get('placa_veiculo'):
            filtros_text.append(f"Placa: {filtros['placa_veiculo']}")
        
        if filtros_text:
            elements.append(Paragraph(f"<b>Filtros Aplicados:</b> {' | '.join(filtros_text)}", styles['Normal']))
            elements.append(Spacer(1, 0.5*cm))
        
        # Buscar MDF-es do banco
        sql = """
            SELECT 
                numero,
                data_emissao,
                placa_veiculo,
                condutor_nome,
                uf_origem,
                uf_destino,
                peso_bruto_total,
                quantidade_documentos,
                status
            FROM manifesto_eletronico
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(data_emissao) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(data_emissao) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND status = %s"
            params.append(filtros['status'])
        if filtros.get('placa_veiculo'):
            sql += " AND placa_veiculo LIKE %s"
            params.append(f"%{filtros['placa_veiculo']}%")
        
        sql += " ORDER BY data_emissao DESC LIMIT 100"
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            mdfes = cursor.fetchall()
        
        # Criar tabela
        data = [['Número', 'Data', 'Placa', 'Condutor', 'Percurso', 'Peso (kg)', 'Qtd Docs', 'Status']]
        total_peso = 0
        total_docs = 0
        
        for mdfe in mdfes:
            numero, data_emissao, placa, condutor, uf_origem, uf_destino, peso, qtd_docs, status = mdfe
            data_formatada = data_emissao.strftime('%d/%m/%Y') if data_emissao else '-'
            peso_num = float(peso) if peso else 0
            percurso = f"{uf_origem or '?'} → {uf_destino or '?'}"
            
            data.append([
                str(numero or '-'),
                data_formatada,
                placa or 'N/A',
                (condutor or 'N/A')[:15],
                percurso,
                f'{peso_num:,.2f}',
                str(qtd_docs or 0),
                (status or 'Pendente')[:12]
            ])
            total_peso += peso_num
            total_docs += int(qtd_docs) if qtd_docs else 0
        
        if len(data) == 1:
            data.append(['-', '-', '-', 'Nenhum MDF-e encontrado', '-', '-', '-', '-'])
        
        table = Table(data, colWidths=[2*cm, 2*cm, 2*cm, 3*cm, 2.5*cm, 2*cm, 1.5*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6a1b9a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Totais
        elements.append(Paragraph(f"<b>Total de MDF-es:</b> {len(mdfes)}", styles['Normal']))
        elements.append(Paragraph(f"<b>Peso Total:</b> {total_peso:,.2f} kg", styles['Normal']))
        elements.append(Paragraph(f"<b>Total de Documentos:</b> {total_docs}", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        return HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    def gerar_relatorio_mdfe_excel(self, filtros):
        """Gera relatório de MDF-e em Excel"""
        if not EXCEL_AVAILABLE:
            return Response({'error': 'Biblioteca openpyxl não disponível'}, status=500)
        
        from django.db import connection
        
        # Buscar dados
        sql = """
            SELECT 
                numero, data_emissao, placa_veiculo, condutor_nome,
                uf_origem, uf_destino, peso_bruto_total, quantidade_documentos, status, chave_acesso
            FROM manifesto_eletronico
            WHERE 1=1
        """
        params = []
        
        if filtros.get('dataInicio'):
            sql += " AND DATE(data_emissao) >= %s"
            params.append(filtros['dataInicio'])
        if filtros.get('dataFim'):
            sql += " AND DATE(data_emissao) <= %s"
            params.append(filtros['dataFim'])
        if filtros.get('status') and filtros['status'] != 'todos':
            sql += " AND status = %s"
            params.append(filtros['status'])
        
        sql += " ORDER BY data_emissao DESC"
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            mdfes = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'MDF-e'
        
        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = 'RELATÓRIO DE MDF-e'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Cabeçalhos
        headers = ['Número', 'Data Emissão', 'Placa Veículo', 'Condutor', 'UF Origem', 'UF Destino', 'Peso Total (kg)', 'Qtd Docs', 'Status', 'Chave Acesso']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='6a1b9a', end_color='6a1b9a', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        row = 4
        total_peso = 0
        total_docs = 0
        for mdfe in mdfes:
            numero, data_emissao, placa, condutor, uf_orig, uf_dest, peso, qtd_docs, status, chave = mdfe
            ws.cell(row=row, column=1, value=numero)
            ws.cell(row=row, column=2, value=data_emissao.strftime('%d/%m/%Y') if data_emissao else '')
            ws.cell(row=row, column=3, value=placa)
            ws.cell(row=row, column=4, value=condutor)
            ws.cell(row=row, column=5, value=uf_orig)
            ws.cell(row=row, column=6, value=uf_dest)
            ws.cell(row=row, column=7, value=float(peso) if peso else 0)
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=qtd_docs)
            ws.cell(row=row, column=9, value=status)
            ws.cell(row=row, column=10, value=chave)
            total_peso += float(peso) if peso else 0
            total_docs += int(qtd_docs) if qtd_docs else 0
            row += 1
        
        # Totais
        row += 1
        ws.cell(row=row, column=6, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=row, column=7, value=total_peso).font = Font(bold=True)
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=total_docs).font = Font(bold=True)
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 50
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=mdfe_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response
