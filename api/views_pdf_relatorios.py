"""
Views para Geração de Relatórios em PDF
Endpoints que geram e retornam PDFs de relatórios fiscais e gerenciais

Autor: Bruno (Sistema Gerencial)
Data: 17/03/2026
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
import datetime
import logging
from django.utils import timezone

from api.services_pdf_fiscal import pdf_fiscal_service

logger = logging.getLogger(__name__)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_cte_condutores(request):
    """
    Retorna lista de condutores distintos cadastrados nos CT-e.
    GET /api/relatorios/cte/condutores/
    """
    try:
        from cte.models import ConhecimentoTransporte
        condutores = (
            ConhecimentoTransporte.objects
            .exclude(condutor_nome__isnull=True)
            .exclude(condutor_nome='')
            .values('condutor_nome', 'condutor_cpf')
            .distinct()
            .order_by('condutor_nome')
        )
        return Response([
            {'nome': c['condutor_nome'], 'cpf': c['condutor_cpf'] or ''}
            for c in condutores
        ])
    except Exception as e:
        logger.error(f"Erro ao listar condutores: {e}", exc_info=True)
        return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_cte_json(request):
    """
    Retorna dados de CT-e em JSON para exibição na tabela do frontend.

    GET /api/relatorios/cte/dados/?data_inicio=2026-01-01&data_fim=2026-03-31
    Params opcionais: destinatario, remetente, status, numero_cte, chave_acesso, tipo_servico
    """
    try:
        from cte.models import ConhecimentoTransporte
        from django.db.models import Q, Sum

        qs = ConhecimentoTransporte.objects.select_related('remetente', 'destinatario').order_by('-data_emissao')

        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        destinatario_id = request.query_params.get('destinatario')
        remetente_id = request.query_params.get('remetente')
        status_filtro = request.query_params.get('status', 'todos')
        numero_cte = request.query_params.get('numero_cte')
        chave_acesso = request.query_params.get('chave_acesso')
        tipo_servico = request.query_params.get('tipo_servico', 'todos')
        condutor = request.query_params.get('condutor')

        if data_inicio_str:
            dt_inicio = timezone.make_aware(datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d'))
            qs = qs.filter(data_emissao__gte=dt_inicio)
        if data_fim_str:
            dt_fim = timezone.make_aware(datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
            qs = qs.filter(data_emissao__lte=dt_fim)
        if destinatario_id:
            qs = qs.filter(destinatario_id=destinatario_id)
        if remetente_id:
            qs = qs.filter(remetente_id=remetente_id)
        if status_filtro and status_filtro != 'todos':
            qs = qs.filter(status_cte__iexact=status_filtro)
        if numero_cte:
            qs = qs.filter(numero_cte=numero_cte)
        if chave_acesso:
            qs = qs.filter(chave_cte__icontains=chave_acesso)
        if tipo_servico and tipo_servico != 'todos':
            qs = qs.filter(tipo_servico=tipo_servico)
        if condutor:
            qs = qs.filter(condutor_nome__iexact=condutor)

        TIPO_SERVICO_MAP = {0: 'Normal', 1: 'Subcontratação', 2: 'Redespacho', 3: 'Redespacho Intermediário', 4: 'Multimodal'}

        # Busca nomes de remetente/destinatário para filtros_aplicados
        remetente_nome = None
        destinatario_nome = None
        if remetente_id:
            from api.models import Cliente
            try:
                remetente_nome = Cliente.objects.get(pk=remetente_id).nome_razao_social
            except Cliente.DoesNotExist:
                pass
        if destinatario_id:
            from api.models import Cliente
            try:
                destinatario_nome = Cliente.objects.get(pk=destinatario_id).nome_razao_social
            except Cliente.DoesNotExist:
                pass

        ctes = []
        for c in qs[:500]:
            ctes.append({
                'id': c.id_cte,
                'numero': c.numero_cte,
                'serie': c.serie_cte,
                'data_emissao': c.data_emissao.strftime('%d/%m/%Y %H:%M') if c.data_emissao else None,
                'chave_acesso': c.chave_cte,
                'protocolo': c.protocolo_cte,
                'remetente': c.remetente.nome_razao_social if c.remetente else None,
                'remetente_id': c.remetente_id,
                'destinatario': c.destinatario.nome_razao_social if c.destinatario else None,
                'destinatario_id': c.destinatario_id,
                'origem': f"{c.cidade_origem_nome or ''}/{c.cidade_origem_uf or ''}".strip('/'),
                'destino': f"{c.cidade_destino_nome or ''}/{c.cidade_destino_uf or ''}".strip('/'),
                'placa': c.placa_veiculo,
                'veiculo_uf': c.veiculo_uf,
                'veiculo_renavam': c.veiculo_renavam,
                'rntrc': c.rntrc,
                'condutor': c.condutor_nome,
                'condutor_cpf': c.condutor_cpf,
                'produto': c.produto_predominante,
                'valor_total': float(c.valor_total_servico or 0),
                'tipo_servico': TIPO_SERVICO_MAP.get(c.tipo_servico, 'Normal'),
                'status': c.status_cte,
            })

        totais = qs.aggregate(valor_total=Sum('valor_total_servico'))
        return Response({
            'ctes': ctes,
            'totais': {
                'quantidade': len(ctes),
                'valor_total': float(totais['valor_total'] or 0),
                'autorizados': qs.filter(status_cte='AUTORIZADO').count(),
                'cancelados': qs.filter(status_cte='CANCELADO').count(),
            },
            'filtros_aplicados': [
                f for f in [
                    f"Período: {data_inicio_str} a {data_fim_str}" if data_inicio_str and data_fim_str else None,
                    f"Remetente: {remetente_nome}" if remetente_nome else None,
                    f"Destinatário: {destinatario_nome}" if destinatario_nome else None,
                    f"Status: {status_filtro}" if status_filtro != 'todos' else None,
                    f"Nº CT-e: {numero_cte}" if numero_cte else None,
                    f"Chave: ...{chave_acesso[-12:]}" if chave_acesso else None,
                    f"Tipo Serviço: {TIPO_SERVICO_MAP.get(int(tipo_servico), tipo_servico)}" if tipo_servico != 'todos' else None,
                    f"Condutor: {condutor}" if condutor else None,
                ] if f
            ]
        })
    except Exception as e:
        logger.error(f"Erro ao buscar CT-es: {e}", exc_info=True)
        return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_cte_excel(request):
    """
    Exporta CT-e em Excel (.xlsx)

    GET /api/relatorios/cte/excel/?data_inicio=2026-01-01&data_fim=2026-03-31
    Params opcionais: destinatario, remetente, status, numero_cte, chave_acesso, tipo_servico
    """
    try:
        from cte.models import ConhecimentoTransporte
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        qs = ConhecimentoTransporte.objects.select_related('remetente', 'destinatario').order_by('-data_emissao')

        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        destinatario_id = request.query_params.get('destinatario')
        remetente_id = request.query_params.get('remetente')
        status_filtro = request.query_params.get('status', 'todos')
        numero_cte = request.query_params.get('numero_cte')
        chave_acesso = request.query_params.get('chave_acesso')
        tipo_servico = request.query_params.get('tipo_servico', 'todos')
        condutor = request.query_params.get('condutor')

        if data_inicio_str:
            dt_inicio = timezone.make_aware(datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d'))
            qs = qs.filter(data_emissao__gte=dt_inicio)
        if data_fim_str:
            dt_fim = timezone.make_aware(datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
            qs = qs.filter(data_emissao__lte=dt_fim)
        if destinatario_id:
            qs = qs.filter(destinatario_id=destinatario_id)
        if remetente_id:
            qs = qs.filter(remetente_id=remetente_id)
        if status_filtro and status_filtro != 'todos':
            qs = qs.filter(status_cte__iexact=status_filtro)
        if numero_cte:
            qs = qs.filter(numero_cte=numero_cte)
        if chave_acesso:
            qs = qs.filter(chave_cte__icontains=chave_acesso)
        if tipo_servico and tipo_servico != 'todos':
            qs = qs.filter(tipo_servico=tipo_servico)
        if condutor:
            qs = qs.filter(condutor_nome__iexact=condutor)

        TIPO_SERVICO_MAP = {0: 'Normal', 1: 'Subcontratação', 2: 'Redespacho', 3: 'Redespacho Intermediário', 4: 'Multimodal'}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatório CT-e'

        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', vertical='center')

        headers = ['Nº CT-e', 'Série', 'Data Emissão', 'Remetente', 'Destinatário',
                   'Origem', 'Destino', 'Placa', 'UF Veíc.', 'RENAVAM', 'RNTRC',
                   'Condutor', 'CPF Condutor', 'Produto',
                   'Valor (R$)', 'Tipo Serviço', 'Status', 'Chave Acesso']
        col_widths = [10, 8, 18, 35, 35, 20, 20, 12, 8, 20, 12, 30, 16, 30, 14, 22, 14, 50]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

        for row_idx, c in enumerate(qs[:5000], 2):
            ws.append([
                c.numero_cte,
                c.serie_cte,
                c.data_emissao.strftime('%d/%m/%Y %H:%M') if c.data_emissao else '',
                c.remetente.nome_razao_social if c.remetente else '',
                c.destinatario.nome_razao_social if c.destinatario else '',
                f"{c.cidade_origem_nome or ''}/{c.cidade_origem_uf or ''}".strip('/'),
                f"{c.cidade_destino_nome or ''}/{c.cidade_destino_uf or ''}".strip('/'),
                c.placa_veiculo or '',
                c.veiculo_uf or '',
                c.veiculo_renavam or '',
                c.rntrc or '',
                c.condutor_nome or '',
                c.condutor_cpf or '',
                c.produto_predominante or '',
                float(c.valor_total_servico or 0),
                TIPO_SERVICO_MAP.get(c.tipo_servico, 'Normal'),
                c.status_cte or '',
                c.chave_cte or '',
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        periodo = f"{data_inicio_str or 'inicio'}_{data_fim_str or 'fim'}"
        filename = f"Relatorio_CTe_{periodo}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"Erro ao gerar Excel CT-e: {e}", exc_info=True)
        return Response({'erro': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_cte_pdf(request):
    """
    Gera PDF do relatório de CT-e
    
    GET /api/relatorios/cte/pdf/?data_inicio=2026-02-01&data_fim=2026-03-17
    
    Query params:
        data_inicio: Data inicial (YYYY-MM-DD)
        data_fim: Data final (YYYY-MM-DD)
    """
    try:
        # Extrai parâmetros
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        
        # Valida parâmetros
        if not data_inicio_str or not data_fim_str:
            return Response({
                'erro': 'Parâmetros obrigatórios: data_inicio e data_fim (formato YYYY-MM-DD)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Converte strings para date
        try:
            data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'erro': 'Formato de data inválido. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Valida intervalo
        if data_inicio > data_fim:
            return Response({
                'erro': 'data_inicio não pode ser maior que data_fim'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"Gerando relatório CT-e: {data_inicio} a {data_fim} (usuário: {request.user.username})")
        
        # Gera o PDF
        pdf_buffer = pdf_fiscal_service.gerar_pdf_cte(data_inicio, data_fim)
        
        # Prepara resposta
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f'Relatorio_CTe_{data_inicio.strftime("%Y%m%d")}_{data_fim.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Relatório CT-e gerado com sucesso: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório CT-e: {e}", exc_info=True)
        return Response({
            'erro': f'Erro ao gerar relatório: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_vendas_operacao_pdf(request):
    """
    Gera PDF do relatório de vendas por operação fiscal
    
    GET /api/relatorios/vendas-operacao/pdf/?data_inicio=2026-02-01&data_fim=2026-03-17
    
    Query params:
        data_inicio: Data inicial (YYYY-MM-DD)
        data_fim: Data final (YYYY-MM-DD)
    """
    try:
        # Extrai parâmetros
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        
        # Valida parâmetros
        if not data_inicio_str or not data_fim_str:
            return Response({
                'erro': 'Parâmetros obrigatórios: data_inicio e data_fim (formato YYYY-MM-DD)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Converte strings para date
        try:
            data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'erro': 'Formato de data inválido. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Valida intervalo
        if data_inicio > data_fim:
            return Response({
                'erro': 'data_inicio não pode ser maior que data_fim'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        logger.info(f"Gerando relatório vendas por operação: {data_inicio} a {data_fim} (usuário: {request.user.username})")
        
        # Gera o PDF
        pdf_buffer = pdf_fiscal_service.gerar_pdf_vendas_operacao(data_inicio, data_fim)
        
        # Prepara resposta
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f'Vendas_por_Operacao_{data_inicio.strftime("%Y%m%d")}_{data_fim.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Relatório vendas por operação gerado com sucesso: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório vendas por operação: {e}", exc_info=True)
        return Response({
            'erro': f'Erro ao gerar relatório: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_vendas_pdf(request):
    """
    Gera PDF do relatório completo de vendas com múltiplos resumos
    
    GET /api/relatorios/vendas/pdf/?data_inicio=2026-02-01&data_fim=2026-03-17&device=mobile
    
    Query params:
        data_inicio: Data inicial (YYYY-MM-DD)
        data_fim: Data final (YYYY-MM-DD)
        cliente: ID do cliente (opcional)
        vendedor: ID do vendedor (opcional)
        operacao: ID da operação (opcional)
        status: Status da venda (opcional: todos, aberta, faturada, cancelada)
        device: Tipo de dispositivo (opcional: mobile, desktop; padrão: desktop)
    """
    try:
        # Extrai parâmetros
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        cliente_id = request.query_params.get('cliente')
        vendedor_id = request.query_params.get('vendedor')
        operacao_id = request.query_params.get('operacao')
        status_filtro = request.query_params.get('status', 'todos')
        device = request.query_params.get('device', 'desktop')
        resumos_str = request.query_params.get('resumos')
        resumos = resumos_str.split(',') if resumos_str else None
        
        # Valida parâmetros obrigatórios
        if not data_inicio_str or not data_fim_str:
            return Response({
                'erro': 'Parâmetros obrigatórios: data_inicio e data_fim (formato YYYY-MM-DD)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Converte strings para date
        try:
            data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                'erro': 'Formato de data inválido. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Valida intervalo
        if data_inicio > data_fim:
            return Response({
                'erro': 'data_inicio não pode ser maior que data_fim'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        logger.info(f"Gerando relatório de vendas completo: {data_inicio} a {data_fim} (usuário: {request.user.username}, device: {device})")
        
        # Monta filtros
        filtros = {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'cliente_id': cliente_id,
            'vendedor_id': vendedor_id,
            'operacao_id': operacao_id,
            'status': status_filtro,
            'resumos': resumos,
            'device': device
        }
        
        # Gera o PDF
        pdf_buffer = pdf_fiscal_service.gerar_pdf_vendas_completo(filtros)
        
        # Prepara resposta
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f'Relatorio_Vendas_{data_inicio.strftime("%Y%m%d")}_{data_fim.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Relatório de vendas completo gerado com sucesso: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório de vendas: {e}", exc_info=True)
        return Response({
            'erro': f'Erro ao gerar relatório: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_estoque_pdf(request):
    """
    Gera PDF do relatório de estoque (placeholder - implementar conforme necessidade)
    
    GET /api/relatorios/estoque/pdf/
    """
    return Response({
        'mensagem': 'Relatório de estoque ainda não implementado.'
    }, status=status.HTTP_501_NOT_IMPLEMENTED)


from api.services_pdf_gerencial import PDFGerencialService


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_financeiro_pdf(request):
    """
    Gera PDF do relatório financeiro com filtros opcionais.
    
    GET /api/relatorios/financeiro/pdf/?data_inicio=...&data_fim=...
    
    Filtros (opcionais):
    - centro_custo_id: ID do Centro de Custo
    - forma_pagamento: Nome da Forma de Pagamento
    - conta_baixa_id: ID da Conta Bancária de baixa
    - conta_lancamento_id: ID da Conta Bancária de lançamento/cobrança
    """
    try:
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        
        if not data_inicio_str or not data_fim_str:
            return Response({'erro': 'Parâmetros data_inicio e data_fim são obrigatórios.'}, status=status.HTTP_400_BAD_REQUEST)
        
        data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date()

        if data_inicio > data_fim:
            return Response({'erro': 'Data de início não pode ser maior que a data de fim.'}, status=status.HTTP_400_BAD_REQUEST)

        filtros = {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'centro_custo_id': request.query_params.get('centro_custo_id'),
            'forma_pagamento': request.query_params.get('forma_pagamento'),
            'conta_baixa_id': request.query_params.get('conta_baixa_id'),
            'conta_lancamento_id': request.query_params.get('conta_lancamento_id'),
            'cliente': request.query_params.get('cliente'),
            'fornecedor': request.query_params.get('fornecedor'),
            'tipo': request.query_params.get('tipo'),
            'status': request.query_params.get('status'),
        }

        logger.info(f"Gerando relatório financeiro: {data_inicio} a {data_fim} com filtros {filtros}")
        
        pdf_buffer = PDFGerencialService.gerar_pdf_financeiro(filtros)
        
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f'Relatorio_Financeiro_{data_inicio.strftime("%Y%m%d")}_{data_fim.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Relatório financeiro gerado: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório financeiro: {e}", exc_info=True)
        return Response({'erro': f'Erro ao gerar relatório: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_conferencia_pdf(request):
    """
    Gera PDF do Relatório de Conferência.

    Cruza Vendas x Contas a Receber x Valores Recebidos x Bancário e exibe divergências.

    GET /api/relatorios/conferencia/pdf/?data_inicio=2026-03-01&data_fim=2026-03-23
    """
    try:
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')

        if not data_inicio_str or not data_fim_str:
            return Response(
                {'erro': 'Parâmetros obrigatórios: data_inicio e data_fim (YYYY-MM-DD)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'erro': 'Formato de data inválido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if data_inicio > data_fim:
            return Response(
                {'erro': 'data_inicio não pode ser maior que data_fim'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        logger.info(
            f"Gerando relatório de conferência: {data_inicio} a {data_fim} "
            f"(usuário: {request.user.username})"
        )

        filtros = {'data_inicio': data_inicio, 'data_fim': data_fim}
        pdf_buffer = pdf_fiscal_service.gerar_pdf_conferencia(filtros)

        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = (
            f'Conferencia_{data_inicio.strftime("%Y%m%d")}_{data_fim.strftime("%Y%m%d")}.pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        logger.info(f'Relatório de conferência gerado: {filename}')
        return response

    except Exception as e:
        logger.error(f'Erro ao gerar relatório de conferência: {e}', exc_info=True)
        return Response(
            {'erro': f'Erro ao gerar relatório: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

